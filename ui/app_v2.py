from __future__ import annotations

import base64
import html
import json
import sys
import tempfile
import warnings
import zipfile
from io import BytesIO
from collections import Counter, defaultdict
from dataclasses import replace
from datetime import date, datetime
from pathlib import Path

import pandas as pd  # type: ignore
import streamlit as st  # type: ignore


# ============================================================
# Project and imports
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


from core.colors import RED
from core.engine import run_rules
from core.findings import Severity, Status
from core.adobe_tag_policy_reconciliation import (
    reconcile_adobe_tag_policy,
)
from core.tag_coverage_reconciliation import reconcile_tag_coverage
from core.adobe_pixel_reconciliation import (
    load_adobe_vendor_rows,
    reconcile_adobe_pixels,
    save_adobe_vendor_rows,
)
from core.dv_reconciliation import reconcile_dv_tags
from core.dv_omni_reconciliation import reconcile_dv_omni
from core.pixel_reconciliation import (
    F_1X1,
    F_DISPLAY,
    F_VIDEO,
    IMPRESSION,
    SURVEY,
    load_vendor_rows,
    reconcile_pixels,
    save_vendor_rows,
)
from core.default_ads import reconcile_default_ads
from core.tag_inventory import (
    TagInventory,
    build_tag_inventory_from_results,
)
from core.matching import match
from core.normalize import norm_compare, norm_dims, site_names_match
from core.tag_matching import match_tags
from core.pdf_report import ReportMeta, build_pdf_report
from core.excel_report import build_excel_report
from parsers.dv_tags import parse_dv_tags
from parsers.innovid_export import parse_innovid_export
from parsers.innovid_tags import parse_innovid_tags
from parsers.ts_parser import detect_profile, parse_ts
from rules import tags as tag_rules
from rules import adobe_pixels
from rules import adobe_tag_policy
from rules import tag_coverage as tag_coverage_rules
from rules import dv_tags as dv_rules
from rules import dv_omni as dv_omni_rules
from rules import pixels as pixel_rules
from rules import defaults as default_rules


warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    module="openpyxl",
)


# ============================================================
# Visual constants
# ============================================================

PROFILE_LABELS = {
    "AUTO": "Auto-detect",
    "adobe_variante_a": (
        "Adobe · Decision Tree Implementation"
    ),
    "adobe_variante_b": (
        "Adobe · Direct & Site-Served Implementation"
    ),
    "wpp_standard": (
        "WPP Media · Standard Trafficking"
    ),
}
def professional_profile_name(profile_name: str) -> str:
    """Corporate-facing name that avoids exposing internal technical keys."""

    return PROFILE_LABELS.get(
        profile_name,
        profile_name.replace("_", " ").title(),
    )


STATUS_PRIORITY = {
    "FAIL": 5,
    "REVIEW": 4,
    "NOT_VERIFIED": 3,
    "INFO": 2,
    "PASS": 1,
}


STATUS_ICON = {
    "PASS": "✅",
    "FAIL": "❌",
    "REVIEW": "⚠️",
    "NOT_VERIFIED": "◻️",
    "INFO": "ℹ️",
}


STATUS_LABEL = {
    "PASS": "Passed",
    "FAIL": "Failed",
    "REVIEW": "Review",
    "NOT_VERIFIED": "Not Verified",
    "INFO": "Info",
}


STATUS_COLOR = {
    "PASS": "#0FA97C",
    "FAIL": "#DC2626",
    "REVIEW": "#D97706",
    "NOT_VERIFIED": "#64748B",
    "INFO": "#0EA5C4",
}


VERDICT_LABELS = {
    "PASSED": "PASSED",
    "FAILED": "REQUIRES CORRECTION",
    "BLOCKED": "BLOCKED",
    "NEEDS_REVIEW": "REVIEW REQUIRED",
    "NO_CHECKS": "NO CHECKS RUN",
}


VERDICT_COLORS = {
    "PASSED": "#0FA97C",
    "FAILED": "#DC2626",
    "BLOCKED": "#991B1B",
    "NEEDS_REVIEW": "#D97706",
    "NO_CHECKS": "#64748B",
}


# ============================================================
# File helpers
# ============================================================

class _RestoredUpload:
    """
    Stand-in for Streamlit's UploadedFile, built from bytes stored in
    a QA2 session bundle -- exposes just the surface the app actually
    touches (.name, .size, .getbuffer(), .getvalue()) so a restored
    file can be passed anywhere a live upload is expected.
    """

    def __init__(self, name: str, data: bytes):
        self.name = name
        self.size = len(data)
        self._data = data

    def getbuffer(self) -> bytes:
        return self._data

    def getvalue(self) -> bytes:
        return self._data


def save_upload(
    uploaded_file,
    directory: Path,
    prefix: str,
) -> Path:
    """
    Saves an UploadedFile to the temporary folder.

    The prefix avoids collisions when two files share the same name.
    """
    safe_name = Path(uploaded_file.name).name
    destination = directory / f"{prefix}{safe_name}"
    destination.write_bytes(uploaded_file.getbuffer())
    return destination


def peek_campaign_name(uploaded_file) -> str:
    """
    Reads just the Campaign Name cell from a TS's Campaign Information
    sheet, without running the full parse_ts pipeline (colors, scope,
    placements...). Used to prefill the Implementation Record's
    Campaign field, which renders before the TS is fully parsed.

    Cheap and best-effort: any failure returns "" so a malformed or
    unusual file never blocks the uploader.
    """
    try:
        from openpyxl import load_workbook

        from core.normalize import norm_key

        wb = load_workbook(
            BytesIO(uploaded_file.getbuffer()),
            read_only=True,
            data_only=True,
        )
        try:
            sheet = next(
                (
                    ws for ws in wb.worksheets
                    if norm_key(ws.title) in ("campaigninformation", "campaigninfo")
                ),
                None,
            )
            if sheet is None:
                return ""

            for row in sheet.iter_rows(min_row=1, max_row=45):
                for idx, cell in enumerate(row):
                    if norm_key(str(cell.value or "").rstrip(":")) != "campaignname":
                        continue
                    for next_cell in row[idx + 1:]:
                        if next_cell.value not in (None, ""):
                            return str(next_cell.value).strip()
            return ""
        finally:
            wb.close()
    except Exception:
        return ""


def anomaly_rows(
    source_name: str,
    result,
) -> list[dict]:
    rows = []

    for anomaly in getattr(result, "anomalies", []):
        rows.append(
            {
                "Source": source_name,
                "Severity": anomaly.severity,
                "Code": anomaly.code,
                "Message": anomaly.message,
                "Reference": (
                    str(anomaly.ref)
                    if getattr(anomaly, "ref", None)
                    else ""
                ),
            }
        )

    return rows


def result_is_fatal(result) -> bool:
    return any(
        anomaly.severity == "FATAL"
        for anomaly in getattr(result, "anomalies", [])
    )


# ============================================================
# Findings helpers
# ============================================================

def placement_findings(
    placement_id: str,
    findings_buffer,
):
    return [
        finding
        for finding in findings_buffer.findings
        if finding.placement_id == placement_id
    ]


def placement_status(
    placement_id: str,
    findings_buffer,
) -> str:
    statuses = [
        finding.status.value
        for finding in placement_findings(
            placement_id,
            findings_buffer,
        )
    ]

    if not statuses:
        return "NOT_VERIFIED"

    return max(
        statuses,
        key=lambda status: STATUS_PRIORITY.get(status, 0),
    )


def findings_dataframe(findings) -> pd.DataFrame:
    rows = []

    for finding in findings:
        rows.append(
            {
                "Status": finding.status.value,
                "Severity": finding.severity.value,
                "Rule": finding.rule_id,
                "Domain": finding.domain.value,
                "Placement ID": finding.placement_id,
                "Placement Name": finding.placement_name,
                "Creative ID": finding.creative_id,
                "Creative Name": finding.creative_name,
                "Message": finding.message,
                "Expected": finding.expected,
                "Found": finding.actual,
                "Reason": finding.reason,
                "Recommended Action": finding.recommended_action,
                "Confidence": finding.confidence.value,
                "Count": finding.count,
            }
        )

    return pd.DataFrame(rows)


def apply_review_overrides(findings, overrides: dict, approved_by: str = ""):
    """
    Turns an approved REVIEW finding into PASS, carrying the QA's
    observation forward as the reason a human can read later.

    `overrides` is {finding_id: {"approved": bool, "note": str}} from
    the sign-off panel. A REVIEW finding not in `overrides` (or not
    approved) passes through unchanged -- only an explicit approval
    flips the status. The finding's identity (rule, placement,
    creative, expected/actual) is untouched, so its finding_id stays
    stable across reruns.
    """
    out = []

    for finding in findings:
        entry = overrides.get(finding.finding_id)

        if not entry or not entry.get("approved") or finding.status.value != "REVIEW":
            out.append(finding)
            continue

        note = entry.get("note", "").strip()
        stamp = f"Approved by {approved_by}" if approved_by else "Approved"
        approval_text = f"{stamp}: {note}" if note else stamp

        out.append(
            replace(
                finding,
                status=Status.PASS,
                severity=Severity.NONE,
                reason=(
                    f"{finding.reason} | {approval_text}"
                    if finding.reason
                    else approval_text
                ),
            )
        )

    return out


def rule_summary_dataframe(findings_buffer) -> pd.DataFrame:
    grouped: dict[str, Counter] = defaultdict(Counter)

    for finding in findings_buffer.findings:
        grouped[finding.rule_id][finding.status.value] += max(
            finding.count,
            1,
        )

    rows = []

    for rule_id in sorted(grouped):
        counts = grouped[rule_id]

        rows.append(
            {
                "Rule": rule_id,
                "PASS": counts.get("PASS", 0),
                "FAIL": counts.get("FAIL", 0),
                "REVIEW": counts.get("REVIEW", 0),
                "NOT_VERIFIED": counts.get(
                    "NOT_VERIFIED",
                    0,
                ),
                "INFO": counts.get("INFO", 0),
                "Total": sum(counts.values()),
            }
        )

    return pd.DataFrame(rows)


# ============================================================
# Visual comparison helpers
# ============================================================

def clean_value(value) -> str:
    if value is None:
        return ""

    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass

    return str(value).strip()


def compare_value(
    expected,
    actual,
    *,
    normalizer=None,
    optional: bool = False,
    fuzzy: bool = False,
) -> str:
    expected_text = clean_value(expected)
    actual_text = clean_value(actual)

    if not expected_text and not actual_text:
        return "INFO" if optional else "NOT_VERIFIED"

    if not expected_text or not actual_text:
        return "INFO" if optional else "NOT_VERIFIED"

    if normalizer is not None:
        expected_comparable = normalizer(expected_text)
        actual_comparable = normalizer(actual_text)
    else:
        expected_comparable = norm_compare(expected_text)
        actual_comparable = norm_compare(actual_text)

    if expected_comparable == actual_comparable:
        return "PASS"

    # El nombre del site se escribe distinto en cada lado y sigue siendo
    # el mismo: "The Trade Desk" en la TS y "FC TradeDesk-DBM" en la
    # plataforma. Basta con que compartan una palabra significativa.
    if fuzzy and site_names_match(expected_comparable, actual_comparable):
        return "PASS"

    return "FAIL"


def comparison_row(
    field_name: str,
    expected,
    actual,
    *,
    normalizer=None,
    optional: bool = False,
    fuzzy: bool = False,
) -> dict:
    return {
        "Validated Field": field_name,
        "Expected in Traffic Sheet": clean_value(expected),
        "Found in Innovid": clean_value(actual),
        "Visual Result": compare_value(
            expected,
            actual,
            normalizer=normalizer,
            fuzzy=fuzzy,
            optional=optional,
        ),
    }


def render_status_badge(status: str) -> None:
    icon = STATUS_ICON.get(status, "")
    label = STATUS_LABEL.get(status, status)
    color = STATUS_COLOR.get(status, "#64748b")

    st.markdown(
        f"""
        <span style="
            display:inline-block;
            padding:5px 12px;
            margin-bottom:10px;
            border-radius:999px;
            background:{color}18;
            border:1px solid {color}40;
            color:{color};
            font-weight:800;
            font-size:0.82rem;
        ">
            {icon} {label}
        </span>
        """,
        unsafe_allow_html=True,
    )


def show_verdict(verdict: str) -> None:
    label = VERDICT_LABELS.get(verdict, verdict)
    color = VERDICT_COLORS.get(verdict, "#64748b")

    st.markdown(
        f"""
        <div class="verdict-card"
             style="border-left:10px solid {color};">
            <div class="verdict-caption">
                OVERALL QA2 RESULT
            </div>
            <div class="verdict-value"
                 style="color:{color};">
                {label}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# Streamlit configuration
# ============================================================

st.set_page_config(
    page_title="Innovid QA2 Automation",
    page_icon="✅",
    layout="wide",
    initial_sidebar_state="expanded",
)


st.markdown(
    """
    <style>
        :root {
            --wpp-indigo: #4B5EEA;
            --wpp-indigo-dark: #2C36A8;
            --wpp-cyan: #17B4DE;
            --wpp-mint: #0FA97C;
            --wpp-ink: #171B2E;
            --wpp-bg: #F5F7FD;
        }

        .stApp {
            background-color: var(--wpp-bg);
        }

        .block-container {
            max-width: 1850px;
            padding-top: 1.2rem;
            padding-bottom: 3rem;
        }

        section[data-testid="stSidebar"] {
            background: linear-gradient(
                180deg,
                #EEF1FE 0%,
                #F5F7FD 55%
            );
            border-right: 1px solid #E2E6FB;
        }

        .wpp-logo-wrap {
            border-radius: 14px;
            overflow: hidden;
            margin-bottom: 18px;
            box-shadow: 0 6px 16px rgba(43, 54, 168, 0.18);
        }

        .qa-header {
            color: white;
            background: linear-gradient(
                120deg,
                var(--wpp-indigo-dark) 0%,
                var(--wpp-indigo) 45%,
                var(--wpp-cyan) 80%,
                var(--wpp-mint) 115%
            );
            padding: 28px 32px;
            border-radius: 18px;
            box-shadow: 0 14px 34px rgba(43, 54, 168, 0.25);
            margin-bottom: 20px;
            position: relative;
            overflow: hidden;
        }

        .qa-header::after {
            content: "";
            position: absolute;
            inset: 0;
            background: radial-gradient(
                circle at 85% -20%,
                rgba(255, 255, 255, 0.35),
                transparent 55%
            );
            pointer-events: none;
        }

        .qa-header h1 {
            margin: 0;
            padding: 0;
            font-size: 2.1rem;
            font-weight: 900;
            letter-spacing: -0.01em;
        }

        .qa-header p {
            margin: 8px 0 0 0;
            opacity: 0.94;
            font-size: 0.97rem;
        }

        .verdict-card {
            background: white;
            border-radius: 16px;
            padding: 20px 24px;
            margin: 10px 0 18px 0;
            box-shadow: 0 6px 20px rgba(23, 27, 46, 0.07);
        }

        .verdict-caption {
            color: #64748b;
            font-size: 0.78rem;
            font-weight: 800;
            letter-spacing: 0.09em;
        }

        .verdict-value {
            margin-top: 4px;
            font-size: 1.95rem;
            font-weight: 900;
        }

        div[data-testid="stMetric"] {
            background: white;
            border: 1px solid #E5E9FA;
            border-radius: 14px;
            padding: 14px;
            box-shadow: 0 3px 12px rgba(23, 27, 46, 0.05);
        }

        div[data-testid="stExpander"] {
            background: white;
            border: 1px solid #E2E6FB;
            border-radius: 13px;
            margin-bottom: 8px;
            overflow: hidden;
            transition: box-shadow 0.15s ease, border-color 0.15s ease;
        }

        div[data-testid="stExpander"]:hover {
            border-color: var(--wpp-indigo);
            box-shadow: 0 4px 14px rgba(75, 94, 234, 0.12);
        }

        div[data-testid="stFileUploader"] {
            background: white;
            border-radius: 13px;
            padding: 6px;
        }

        .upload-help {
            color: #94a3b8;
            font-size: 0.75rem;
            margin-top: -8px;
            margin-bottom: 13px;
            line-height: 1.35;
        }

        .profile-card {
            background: white;
            border: 1px solid #E2E6FB;
            border-radius: 13px;
            padding: 13px 16px;
            margin-bottom: 15px;
            line-height: 1.55;
        }

        .profile-card strong {
            color: var(--wpp-indigo-dark);
        }

        .section-note {
            background: #EEF1FE;
            border-left: 5px solid var(--wpp-indigo);
            border-radius: 10px;
            padding: 11px 14px;
            margin-bottom: 12px;
            color: var(--wpp-indigo-dark);
        }

        .stButton > button {
            width: 100%;
            min-height: 48px;
            border-radius: 11px;
            font-weight: 850;
        }

        .stButton > button[kind="primary"] {
            background: linear-gradient(
                100deg,
                var(--wpp-indigo) 0%,
                var(--wpp-cyan) 130%
            );
            border: none;
            box-shadow: 0 6px 16px rgba(75, 94, 234, 0.3);
            transition: transform 0.12s ease, box-shadow 0.12s ease;
        }

        .stButton > button[kind="primary"]:hover {
            transform: translateY(-1px);
            box-shadow: 0 8px 20px rgba(75, 94, 234, 0.4);
        }

        button[role="tab"][aria-selected="true"] {
            color: var(--wpp-indigo-dark) !important;
        }

        div[data-baseweb="tab-highlight"] {
            background-color: var(--wpp-indigo) !important;
        }
    </style>
    """,
    unsafe_allow_html=True,
)


st.markdown(
    """
    <div class="qa-header">
        <h1>INNOVID QA2 AUTOMATION</h1>
        <p>
            Validation of Traffic Sheet, Placement-Creative View,
            Placement View, and Tag files
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# Editable vendor pixel config (PIX-002: DoubleVerify, Dynata,
# Kantar, Inmarket, DISQO)
# ============================================================

with st.expander("⚙️ Pixels by account (editable) -- WPP"):
    st.caption(
        "Vendor pixel rules PIX-002/DV-001/DV-003 check against "
        "Innovid, for Unilever / Wendy's / BlackRock. DV is split "
        "into its 4 real flavors -- Monitoring, Blocking, Integration, "
        "Omni -- each verified differently; only Monitoring uses a "
        "placement-level pixel, the others are read straight from "
        "their own row by name, so they don't need Host terms/Column/"
        "Format filled in at all."
    )
    st.caption(
        "**Official pixel** is the one field that matters: paste the "
        "vendor's current reference (with its macros, e.g. "
        "[%placementID%]) and QA2 flags REVIEW if what's implemented "
        "in Innovid doesn't match it anymore. Fill it and you can "
        "leave Host terms blank -- the domain is derived from the "
        "pixel automatically. Leave Account blank for a vendor shared "
        "across all three WPP accounts; set it (e.g. \"Wendy's\") for "
        "one that only applies to that account -- pick the matching "
        "Account in the sidebar when you run QA2. Saved to "
        "config/vendor_pixels.json; applies on the next QA2 run for "
        "everyone who pulls this repo after the file is committed."
    )

    _vendor_rows = load_vendor_rows()
    _vendor_df = pd.DataFrame(
        [
            {
                "Account": r.get("account", ""),
                "Vendor": r.get("name", ""),
                "TS terms": ", ".join(r.get("ts_terms", [])),
                "Host terms": ", ".join(r.get("host_terms", [])),
                "Column": r.get("column", IMPRESSION),
                "1x1": F_1X1 in (r.get("formats") or []),
                "Display": F_DISPLAY in (r.get("formats") or []),
                "Video": F_VIDEO in (r.get("formats") or []),
                "Site exceptions": ", ".join(r.get("site_exceptions", [])),
                "Official pixel": r.get("official_pixel", ""),
                "Note": r.get("note", ""),
            }
            for r in _vendor_rows
        ]
    )

    _edited_vendor_df = st.data_editor(
        _vendor_df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Column": st.column_config.SelectboxColumn(
                options=[SURVEY, IMPRESSION]
            ),
        },
        key="qa2_vendor_editor",
    )

    if st.button("Save pixel table", key="qa2_vendor_save"):
        _new_vendor_rows = []
        for _, _row in _edited_vendor_df.iterrows():
            _name = str(_row.get("Vendor") or "").strip()
            if not _name:
                continue
            _formats = []
            if _row.get("1x1"):
                _formats.append(F_1X1)
            if _row.get("Display"):
                _formats.append(F_DISPLAY)
            if _row.get("Video"):
                _formats.append(F_VIDEO)
            _new_vendor_rows.append(
                {
                    "account": str(_row.get("Account") or "").strip(),
                    "name": _name,
                    "ts_terms": [
                        t.strip()
                        for t in str(_row.get("TS terms") or "").split(",")
                        if t.strip()
                    ],
                    "host_terms": [
                        t.strip()
                        for t in str(_row.get("Host terms") or "").split(",")
                        if t.strip()
                    ],
                    "column": _row.get("Column") or IMPRESSION,
                    "formats": _formats,
                    "site_exceptions": [
                        t.strip()
                        for t in str(_row.get("Site exceptions") or "").split(",")
                        if t.strip()
                    ],
                    "official_pixel": str(_row.get("Official pixel") or "").strip(),
                    "note": str(_row.get("Note") or ""),
                }
            )
        save_vendor_rows(_new_vendor_rows)
        st.success(
            f"Saved {len(_new_vendor_rows)} vendor(s) to "
            "config/vendor_pixels.json."
        )

with st.expander("⚙️ Pixels by account (editable) -- Adobe"):
    st.caption(
        "The official pixel Adobe's DISQO/iSpot check (PIX-A01) "
        "compares against once it finds evidence in Innovid or the "
        "tag files. Unlike the WPP table above, this doesn't gate "
        "presence (that's still a broad search across Third Party "
        "Impression, Third Party Survey, Clicktag and the tag files) "
        "-- it only flags when a pixel IS found but doesn't match the "
        "reference below. These are Third Party placements, and the "
        "pixel varies by Adobe campaign (Acrobat, Firefly, STE, PGA, "
        "MLB, Adelaide, Express...), not just by vendor -- leave "
        "Campaign blank for a reference that applies everywhere, or "
        "add one row per campaign with its own pixel and pick the "
        "matching Account / Campaign in the sidebar. Leave the pixel "
        "blank to skip the check for a given row."
    )

    _adobe_vendor_rows = load_adobe_vendor_rows()
    _adobe_vendor_df = pd.DataFrame(
        [
            {
                "Vendor": r.get("name", ""),
                "Campaign": r.get("campaign", ""),
                "Official pixel": r.get("official_pixel", ""),
                "Note": r.get("note", ""),
            }
            for r in _adobe_vendor_rows
        ]
    )

    _edited_adobe_vendor_df = st.data_editor(
        _adobe_vendor_df,
        num_rows="dynamic",
        use_container_width=True,
        key="qa2_adobe_vendor_editor",
    )

    if st.button("Save Adobe pixel table", key="qa2_adobe_vendor_save"):
        _new_adobe_rows = [
            {
                "name": str(_row.get("Vendor") or "").strip(),
                "campaign": str(_row.get("Campaign") or "").strip(),
                "official_pixel": str(_row.get("Official pixel") or "").strip(),
                "note": str(_row.get("Note") or ""),
            }
            for _, _row in _edited_adobe_vendor_df.iterrows()
            if str(_row.get("Vendor") or "").strip()
        ]
        save_adobe_vendor_rows(_new_adobe_rows)
        st.success(
            f"Saved {len(_new_adobe_rows)} vendor(s) to "
            "config/vendor_pixels_adobe.json."
        )


# ============================================================
# Sidebar
# ============================================================

with st.sidebar:
    _logo_path = PROJECT_ROOT / "ui" / "assets" / "wpp-media-logo.png.png"

    if _logo_path.exists():
        _logo_b64 = base64.b64encode(_logo_path.read_bytes()).decode()
        st.markdown(
            f"""
            <div class="wpp-logo-wrap">
                <img src="data:image/png;base64,{_logo_b64}"
                     style="width:100%; display:block;">
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.header("QA2 Files")

    with st.expander("📂 Load a saved session (optional)"):
        st.caption(
            "If the implementer already ran QA2 and used \"Save "
            "session bundle\" below, load that .zip here (same "
            "SharePoint folder as the TS) to skip re-uploading the "
            "same files -- you'll only need to review and check "
            "QA2 Sign-off. You can still swap any file below by "
            "uploading a different one over it."
        )
        uploaded_bundle = st.file_uploader(
            "Session bundle (.zip)",
            type=["zip"],
            key="qa2_session_bundle",
        )

    _restored: dict = {}

    if uploaded_bundle is not None:
        _bundle_source = (uploaded_bundle.name, uploaded_bundle.size)
        _is_new_bundle = (
            st.session_state.get("_session_bundle_source")
            != _bundle_source
        )

        try:
            with zipfile.ZipFile(
                BytesIO(uploaded_bundle.getvalue())
            ) as _zf:
                _manifest = json.loads(
                    _zf.read("manifest.json").decode("utf-8")
                )
                _manifest_files = _manifest.get("files", {})

                def _read_bundle_file(rel_path: str) -> bytes | None:
                    try:
                        return _zf.read(rel_path)
                    except KeyError:
                        return None

                for _slot in ("ts", "pc", "pl", "dv"):
                    _name = _manifest_files.get(_slot)
                    if not _name:
                        continue
                    _data = _read_bundle_file(f"files/{_slot}/{_name}")
                    if _data is not None:
                        _restored[_slot] = _RestoredUpload(_name, _data)

                _restored_tags = []
                for _entry in _manifest_files.get("tags", []):
                    _stored_as = _entry.get("stored_as")
                    _orig_name = _entry.get("name")
                    if not _stored_as or not _orig_name:
                        continue
                    _data = _read_bundle_file(f"files/tags/{_stored_as}")
                    if _data is not None:
                        _restored_tags.append(
                            _RestoredUpload(_orig_name, _data)
                        )
                if _restored_tags:
                    _restored["tags"] = _restored_tags

                if _is_new_bundle:
                    if _manifest.get("profile") in PROFILE_LABELS:
                        st.session_state["qa2_profile_select"] = (
                            _manifest["profile"]
                        )
                    _valid_accounts = {
                        row.get("account", "").strip()
                        for row in load_vendor_rows()
                        if row.get("account", "").strip()
                    } | {
                        row.get("campaign", "").strip()
                        for row in load_adobe_vendor_rows()
                        if row.get("campaign", "").strip()
                    }
                    if _manifest.get("account") in _valid_accounts:
                        st.session_state["qa2_account_select"] = (
                            _manifest["account"]
                        )
                    for _field_key, _meta_key in (
                        ("qa2_record_campaign", "campaign"),
                        ("qa2_record_request_type", "request_type"),
                        ("qa2_record_wrike", "wrike_id"),
                        ("qa2_record_impl_by", "implemented_by"),
                    ):
                        if _manifest.get(_meta_key):
                            st.session_state[_field_key] = (
                                _manifest[_meta_key]
                            )
                    if _manifest.get("implementation_date"):
                        st.session_state["qa2_record_impl_date"] = (
                            date.fromisoformat(
                                _manifest["implementation_date"]
                            )
                        )
                    st.session_state["_session_bundle_source"] = (
                        _bundle_source
                    )

            st.success(
                f"Loaded session bundle -- {len(_restored)} file "
                "slot(s) restored below."
            )
        except Exception as exc:
            st.error(f"Couldn't read this session bundle: {exc}")

    uploaded_ts = st.file_uploader(
        "1. Upload Traffic Sheet",
        type=["xlsx", "xlsm"],
        accept_multiple_files=False,
        key="qa2_ts",
    ) or _restored.get("ts")

    if uploaded_ts is not None:
        # Prefills the Implementation Record's Campaign field from the
        # TS's own Campaign Information sheet. Only re-peeks when the
        # uploaded file actually changes, so it doesn't re-parse on
        # every rerun, and never overwrites a name the user already
        # typed by hand for the current file.
        _campaign_source = (uploaded_ts.name, uploaded_ts.size)
        if st.session_state.get("_campaign_autofill_source") != _campaign_source:
            _peeked_name = peek_campaign_name(uploaded_ts)
            if _peeked_name:
                st.session_state["qa2_record_campaign"] = _peeked_name
            st.session_state["_campaign_autofill_source"] = _campaign_source

    st.markdown(
        """
        <div class="upload-help">
            Source document for scope, placements, and requested changes.
        </div>
        """,
        unsafe_allow_html=True,
    )

    uploaded_pc = st.file_uploader(
        "2. Upload Innovid Placement-Creative View",
        type=["xlsx", "xlsm"],
        accept_multiple_files=False,
        key="qa2_pc",
    ) or _restored.get("pc")

    st.markdown(
        """
        <div class="upload-help">
            Export with Creative_ID, association, status,
            Decision Tree, Clicktag, and Third Party ID.
        </div>
        """,
        unsafe_allow_html=True,
    )

    uploaded_pl = st.file_uploader(
        "3. Upload Innovid Placement View",
        type=["xlsx", "xlsm"],
        accept_multiple_files=False,
        key="qa2_pl",
    ) or _restored.get("pl")

    st.markdown(
        """
        <div class="upload-help">
            Optional. Recommended for 1x1s, pixels,
            and placement-level URL validations.
        </div>
        """,
        unsafe_allow_html=True,
    )

    uploaded_tags = st.file_uploader(
        "4. Upload Tag files",
        type=["xlsx", "xlsm"],
        accept_multiple_files=True,
        key="qa2_tags",
    )
    if not uploaded_tags and _restored.get("tags"):
        uploaded_tags = _restored["tags"]

    st.markdown(
        """
        <div class="upload-help">
            Allows selecting multiple files at once.
            On Windows use Ctrl+click or Shift+click.
        </div>
        """,
        unsafe_allow_html=True,
    )

    uploaded_evidence = st.file_uploader(
        "5. Upload evidence screenshots",
        type=["png", "jpg", "jpeg"],
        accept_multiple_files=True,
        key="qa2_evidence",
    )

    st.markdown(
        """
        <div class="upload-help">
            Optional. Included as an Implementation Evidence
            section in the PDF report.
        </div>
        """,
        unsafe_allow_html=True,
    )

    uploaded_dv = st.file_uploader(
        "6. Upload DV Pinnacle Tags",
        type=["xlsx", "xlsm"],
        accept_multiple_files=False,
        key="qa2_dv",
    ) or _restored.get("dv")

    st.markdown(
        """
        <div class="upload-help">
            Optional. Only needed for placements whose "Vendors /
            Pixels" value mentions DV (DoubleVerify).
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.divider()

    with st.expander("Implementation Record (optional)"):
        st.caption(
            "Included in the PDF report. Leave blank to fill by hand."
        )

        record_campaign = st.text_input(
            "Campaign", key="qa2_record_campaign"
        )
        record_request_type = st.text_input(
            "Request Type",
            placeholder=(
                "e.g. New placements, swaps, creative adds..."
            ),
            key="qa2_record_request_type",
        )
        record_wrike_id = st.text_input(
            "Wrike ID", key="qa2_record_wrike"
        )
        record_implemented_by = st.text_input(
            "Implemented By", key="qa2_record_impl_by"
        )
        record_implementation_date = st.date_input(
            "Implementation Date",
            value=None,
            key="qa2_record_impl_date",
        )
        st.caption(
            "QA2 By / QA2 Date have moved -- fill those in the "
            "QA2 Review section below the results, right where "
            "you approve."
        )
        record_qa3_by = st.text_input(
            "QA3 By", key="qa2_record_qa3_by"
        )
        record_qa3_date = st.date_input(
            "QA3 Date",
            value=None,
            key="qa2_record_qa3_date",
        )
        record_notes = st.text_area(
            "Notes / Callouts",
            placeholder=(
                "Anything worth flagging that doesn't block "
                "implementation but should be on record..."
            ),
            key="qa2_record_notes",
        )

    st.divider()

    _account_options = ["All / unknown"] + sorted(
        {
            row.get("account", "").strip()
            for row in load_vendor_rows()
            if row.get("account", "").strip()
        }
        | {
            row.get("campaign", "").strip()
            for row in load_adobe_vendor_rows()
            if row.get("campaign", "").strip()
        }
    )
    selected_account = st.selectbox(
        "Account / Campaign",
        options=_account_options,
        index=0,
        key="qa2_account_select",
        help=(
            "Some vendor pixel rules only apply to one account or "
            "campaign (e.g. Inmarket and DISQO are Wendy's-only; "
            "Adobe's official pixels vary per campaign -- Acrobat, "
            "Firefly, STE...). Pick the one this Traffic Sheet "
            "belongs to so PIX-002/PIX-A01 apply the right rows from "
            "the Pixels by account panels. Add new options there by "
            "filling Account (WPP) or Campaign (Adobe) on a row."
        ),
    )
    if selected_account == "All / unknown":
        selected_account = ""

    selected_profile = st.selectbox(
        "Traffic Sheet Profile",
        options=list(PROFILE_LABELS),
        format_func=lambda value: PROFILE_LABELS[value],
        index=0,
        key="qa2_profile_select",
    )

    analyze_button = st.button(
        "Run QA2",
        type="primary",
        use_container_width=True,
    )

    # st.button() only returns True on the single rerun triggered by
    # the click itself -- on every later rerun (e.g. changing a filter
    # in the results tabs) it goes back to False. Without persisting
    # this in session_state, touching any filter after running QA2
    # would drop back to the landing screen and appear to "reset"
    # the whole app, discarding the analysis.
    if "qa2_has_run" not in st.session_state:
        st.session_state.qa2_has_run = False

    if analyze_button:
        st.session_state.qa2_has_run = True

    st.divider()

    _can_bundle = uploaded_ts is not None and uploaded_pc is not None

    if _can_bundle:
        _bundle_buffer = BytesIO()
        with zipfile.ZipFile(
            _bundle_buffer, "w", zipfile.ZIP_DEFLATED
        ) as _zf:
            _files_manifest: dict = {}

            _ts_name = Path(uploaded_ts.name).name
            _zf.writestr(f"files/ts/{_ts_name}", uploaded_ts.getbuffer())
            _files_manifest["ts"] = _ts_name

            _pc_name = Path(uploaded_pc.name).name
            _zf.writestr(f"files/pc/{_pc_name}", uploaded_pc.getbuffer())
            _files_manifest["pc"] = _pc_name

            if uploaded_pl is not None:
                _pl_name = Path(uploaded_pl.name).name
                _zf.writestr(
                    f"files/pl/{_pl_name}", uploaded_pl.getbuffer()
                )
                _files_manifest["pl"] = _pl_name

            if uploaded_dv is not None:
                _dv_name = Path(uploaded_dv.name).name
                _zf.writestr(
                    f"files/dv/{_dv_name}", uploaded_dv.getbuffer()
                )
                _files_manifest["dv"] = _dv_name

            _tag_entries = []
            for _index, _tag_file in enumerate(uploaded_tags or [], start=1):
                _stored_as = f"{_index}_{Path(_tag_file.name).name}"
                _zf.writestr(
                    f"files/tags/{_stored_as}", _tag_file.getbuffer()
                )
                _tag_entries.append(
                    {
                        "name": Path(_tag_file.name).name,
                        "stored_as": _stored_as,
                    }
                )
            if _tag_entries:
                _files_manifest["tags"] = _tag_entries

            _manifest_out = {
                "profile": selected_profile,
                "account": selected_account,
                "campaign": record_campaign,
                "request_type": record_request_type,
                "wrike_id": record_wrike_id,
                "implemented_by": record_implemented_by,
                "implementation_date": (
                    record_implementation_date.isoformat()
                    if record_implementation_date else ""
                ),
                "saved_at": datetime.now().isoformat(timespec="seconds"),
                "files": _files_manifest,
            }
            _zf.writestr(
                "manifest.json", json.dumps(_manifest_out, indent=2)
            )

        st.download_button(
            "💾 Save session bundle (for QA2)",
            data=_bundle_buffer.getvalue(),
            file_name="qa2_session_bundle.zip",
            mime="application/zip",
            use_container_width=True,
            help=(
                "Bundles the uploaded TS, Innovid exports and tag "
                "files plus your Campaign/Profile/Account picks into "
                "one .zip. Drop it in the same SharePoint folder as "
                "the TS -- QA2 loads it from \"Load a saved session\" "
                "above and skips re-uploading everything, straight "
                "to reviewing and checking QA2 Sign-off."
            ),
        )
    else:
        st.caption(
            "Upload the Traffic Sheet and Placement-Creative View "
            "to enable \"Save session bundle\"."
        )


# ============================================================
# Landing screen
# ============================================================

if not st.session_state.qa2_has_run:
    st.info(
        "Upload at least the Traffic Sheet and the "
        "Innovid Placement-Creative View."
    )

    st.subheader("Analysis Flow")

    flow_columns = st.columns(4)

    flow_columns[0].markdown(
        """
        **1. Traffic Sheet**

        Detects account, profile, colors, scope, and worked placements.
        """
    )

    flow_columns[1].markdown(
        """
        **2. Placement-Creative**

        Validates creatives, association, status, Decision Tree, URL, and attribution.
        """
    )

    flow_columns[2].markdown(
        """
        **3. Placement View**

        Adds supporting info for 1x1 placements, URLs, and pixels.
        """
    )

    flow_columns[3].markdown(
        """
        **4. Tags**

        Validates IDs, dimensions, Campaign ID, and delivered content.
        """
    )

    st.stop()


if uploaded_ts is None or uploaded_pc is None:
    st.error(
        "You must upload the Traffic Sheet and the "
        "Innovid Placement-Creative View."
    )
    st.stop()


# ============================================================
# Processing
# ============================================================

with tempfile.TemporaryDirectory(
    prefix="qa2_v2_"
) as temporary_directory:
    temporary_path = Path(temporary_directory)

    ts_path = save_upload(
        uploaded_ts,
        temporary_path,
        "ts_",
    )

    pc_path = save_upload(
        uploaded_pc,
        temporary_path,
        "pc_",
    )

    pl_path = (
        save_upload(
            uploaded_pl,
            temporary_path,
            "pl_",
        )
        if uploaded_pl is not None
        else None
    )

    dv_path = (
        save_upload(
            uploaded_dv,
            temporary_path,
            "dv_",
        )
        if uploaded_dv is not None
        else None
    )

    tag_paths = []

    for index, uploaded_tag in enumerate(
        uploaded_tags or [],
        start=1,
    ):
        tag_paths.append(
            (
                uploaded_tag.name,
                save_upload(
                    uploaded_tag,
                    temporary_path,
                    f"tag_{index}_",
                ),
            )
        )

    try:
        # ----------------------------------------------------
        # Traffic Sheet
        # ----------------------------------------------------

        with st.spinner(
            "Reading and parsing the Traffic Sheet..."
        ):
            detected_profile, detection_evidence = (
                detect_profile(ts_path)
            )

            forced_profile = (
                None
                if selected_profile == "AUTO"
                else selected_profile
            )

            ts_result = parse_ts(
                ts_path,
                profile_name=forced_profile,
            )

        # ----------------------------------------------------
        # Placement-Creative
        # ----------------------------------------------------

        with st.spinner(
            "Reading Innovid Placement-Creative View..."
        ):
            pc_result = parse_innovid_export(pc_path)

        # ----------------------------------------------------
        # Placement View
        # ----------------------------------------------------

        pl_result = None

        if pl_path is not None:
            with st.spinner(
                "Reading Innovid Placement View..."
            ):
                pl_result = parse_innovid_export(pl_path)

        # ----------------------------------------------------
        # Tags
        # ----------------------------------------------------

        tags_results = []

        if tag_paths:
            with st.spinner(
                f"Reading {len(tag_paths)} tag file(s)..."
            ):
                for original_name, tag_path in tag_paths:
                    tags_results.append(
                        (
                            original_name,
                            parse_innovid_tags(tag_path),
                        )
                    )

        # ----------------------------------------------------
        # DV Pinnacle Tags
        # ----------------------------------------------------

        dv_result = None

        if dv_path is not None:
            with st.spinner(
                "Reading DV Pinnacle Tags..."
            ):
                dv_result = parse_dv_tags(dv_path)

        # ----------------------------------------------------
        # File understanding
        # ----------------------------------------------------

        file_rows = [
            {
                "File": uploaded_ts.name,
                "Expected Type": "Traffic Sheet",
                "Detected Type": professional_profile_name(
                    ts_result.profile
                ),
                "Status": (
                    "FATAL"
                    if result_is_fatal(ts_result)
                    else "OK"
                ),
                "Records": (
                    len(ts_result.placements.rows)
                    if ts_result.placements
                    else 0
                ),
            },
            {
                "File": uploaded_pc.name,
                "Expected Type": "Placement-Creative View",
                "Detected Type": (
                    pc_result.level
                    or "Not Recognized"
                ),
                "Status": (
                    "FATAL"
                    if result_is_fatal(pc_result)
                    else "OK"
                ),
                "Records": len(pc_result.rows),
            },
        ]

        if pl_result is not None:
            file_rows.append(
                {
                    "File": uploaded_pl.name,
                    "Expected Type": "Placement View",
                    "Detected Type": (
                        pl_result.level
                        or "Not Recognized"
                    ),
                    "Status": (
                        "FATAL"
                        if result_is_fatal(pl_result)
                        else "OK"
                    ),
                    "Records": len(pl_result.rows),
                }
            )

        for file_name, tags_result in tags_results:
            file_rows.append(
                {
                    "File": file_name,
                    "Expected Type": "Tags",
                    "Detected Type": (
                        f"Tags | sheet {tags_result.sheet}"
                        if tags_result.sheet
                        else "Not Recognized"
                    ),
                    "Status": (
                        "FATAL"
                        if result_is_fatal(tags_result)
                        else "OK"
                    ),
                    "Records": len(tags_result.rows),
                }
            )

        if dv_result is not None:
            file_rows.append(
                {
                    "File": uploaded_dv.name,
                    "Expected Type": "DV Pinnacle Tags",
                    "Detected Type": (
                        f"DV Pinnacle | sheet {dv_result.sheet}"
                        if dv_result.sheet
                        else "Not Recognized"
                    ),
                    "Status": (
                        "FATAL"
                        if result_is_fatal(dv_result)
                        else "OK"
                    ),
                    "Records": len(dv_result.rows),
                }
            )

        files_dataframe = pd.DataFrame(file_rows)

        # ----------------------------------------------------
        # Anomalies
        # ----------------------------------------------------

        all_anomaly_rows = []

        all_anomaly_rows.extend(
            anomaly_rows(
                "Traffic Sheet",
                ts_result,
            )
        )

        all_anomaly_rows.extend(
            anomaly_rows(
                "Placement-Creative View",
                pc_result,
            )
        )

        if pl_result is not None:
            all_anomaly_rows.extend(
                anomaly_rows(
                    "Placement View",
                    pl_result,
                )
            )

        for file_name, tags_result in tags_results:
            all_anomaly_rows.extend(
                anomaly_rows(
                    f"Tags | {file_name}",
                    tags_result,
                )
            )

        if dv_result is not None:
            all_anomaly_rows.extend(
                anomaly_rows(
                    "DV Pinnacle Tags",
                    dv_result,
                )
            )

        # Explicit file-type validations.

        if (
            not result_is_fatal(pc_result)
            and pc_result.level != "placement_creative"
        ):
            all_anomaly_rows.append(
                {
                    "Source": "Placement-Creative View",
                    "Severity": "FATAL",
                    "Code": "UI-WRONG-PC-FILE",
                    "Message": (
                        "The uploaded file was not recognized as a "
                        "Placement-Creative View. It must contain Creative_ID."
                    ),
                    "Reference": uploaded_pc.name,
                }
            )

        if (
            pl_result is not None
            and not result_is_fatal(pl_result)
            and pl_result.level != "placement"
        ):
            all_anomaly_rows.append(
                {
                    "Source": "Placement View",
                    "Severity": "FATAL",
                    "Code": "UI-WRONG-PL-FILE",
                    "Message": (
                        "The uploaded file was not recognized as a "
                        "Placement View. Do not upload a TAGS file here."
                    ),
                    "Reference": uploaded_pl.name,
                }
            )

        fatal_rows = [
            row
            for row in all_anomaly_rows
            if row["Severity"] == "FATAL"
        ]

        if fatal_rows:
            st.error(
                "The analysis was blocked because one or more "
                "documents don't match the expected format."
            )

            st.subheader("File Understanding")

            st.dataframe(
                files_dataframe,
                use_container_width=True,
                hide_index=True,
            )

            st.subheader("Issues Found")

            st.dataframe(
                pd.DataFrame(fatal_rows),
                use_container_width=True,
                hide_index=True,
            )

            st.stop()

        # ----------------------------------------------------
        # Matching and rules
        # ----------------------------------------------------

        with st.spinner(
            "Running QA2 matching and validations..."
        ):
            match_result = match(
                ts_result,
                pc_result,
                pl_result,
            )

            # Run TS vs Innovid rules first.
            findings_buffer = run_rules(match_result)

            tag_matches = []
            tag_inventory = None
            adobe_pixel_result = None

            # Every tag file keeps its independent structural checks.
            for file_name, tags_result in tags_results:
                tag_match_result = match_tags(
                    match_result,
                    tags_result,
                )

                tag_rules.evaluate(
                    tag_match_result,
                    findings_buffer,
                )

                tag_matches.append(
                    (
                        file_name,
                        tags_result,
                        tag_match_result,
                    )
                )

            # Consolidated multi-file Tag Inventory.
            #
            # Structural TAG rules run once per delivered file.
            # Adobe business requirements run once against the
            # consolidated inventory to avoid duplicate findings.
            if tags_results:
                tag_inventory = (
                    build_tag_inventory_from_results(
                        tags_results
                    )
                )

            # PIX-A01 (DISQO/iSpot) applies to both Adobe profiles:
            # Site-Served (adobe_variante_b) and Third Party / Decision
            # Tree (adobe_variante_a). It used to run only for
            # variante_b, so a Decision Tree campaign requiring DISQO
            # never got checked at all.
            #
            # Traffic Sheet defines whether DISQO/iSpot is required.
            # Tags and Innovid Placement View provide evidence.
            if (
                ts_result.profile in ("adobe_variante_a", "adobe_variante_b")
                and tag_inventory is not None
                and pl_result is not None
            ):
                adobe_pixel_result = (
                    reconcile_adobe_pixels(
                        ts_result,
                        pl_result,
                        tag_inventory,
                        selected_account,
                    )
                )

                adobe_pixels.evaluate(
                    adobe_pixel_result,
                    findings_buffer,
                )

            # TAG-012: Adobe tag column policy (Ftrack / No Ftrack /
            # Clicktag / Protected). Only needs the delivered tag
            # files, not the Placement View, so it can run for Adobe
            # even without one uploaded.
            adobe_tag_policy_result = None

            if (
                ts_result.profile in ("adobe_variante_a", "adobe_variante_b")
                and tag_inventory is not None
            ):
                adobe_tag_policy_result = reconcile_adobe_tag_policy(
                    ts_result,
                    tag_inventory,
                )

                adobe_tag_policy.evaluate(
                    adobe_tag_policy_result,
                    findings_buffer,
                )

                # TAG-013: every 1x1 placement with a declared vendor
                # requirement should have a row in the delivered tag
                # file(s) -- coverage, not pixel content.
                tag_coverage_result = reconcile_tag_coverage(
                    ts_result,
                    tag_inventory,
                )

                tag_coverage_rules.evaluate(
                    tag_coverage_result,
                    findings_buffer,
                )

            # DV (DoubleVerify) tag delivery check.
            #
            # Applies to any placement whose Traffic Sheet
            # "Vendors / Pixels" value mentions DV, regardless of
            # profile. Uses an empty inventory when no tag files
            # were uploaded, so delivered-but-unverifiable DV tags
            # surface as REVIEW rather than being silently skipped.
            dv_reconciliation = reconcile_dv_tags(
                ts_result,
                tag_inventory
                if tag_inventory is not None
                else TagInventory(),
                dv_result,
            )

            dv_rules.evaluate(
                dv_reconciliation,
                findings_buffer,
            )

            # DV-003: DV Omni y DV Monitoring/Blocking, que no van por
            # pixel de placement ni por archivo de Pinnacle sino por
            # una columna dentro del archivo de tags de Innovid.
            dv_omni_reconciliation = reconcile_dv_omni(
                ts_result,
                pl_result,
                tag_inventory
                if tag_inventory is not None
                else TagInventory(),
            )

            dv_omni_rules.evaluate(
                dv_omni_reconciliation,
                findings_buffer,
            )

            # Pixeles de vendor declarados en "Vendors / Pixels"
            # (DoubleVerify, Dynata, Kantar) contra lo que Innovid
            # tiene cargado a nivel de placement.
            pixel_reconciliation = reconcile_pixels(
                ts_result,
                pl_result,
                selected_account,
            )

            pixel_rules.evaluate(
                pixel_reconciliation,
                findings_buffer,
            )

            # El default ad es un creativo aparte, el mismo para todos
            # los placements de su dimension: si uno se queda con otro
            # default o con otra landing page, no recibio el swap.
            default_ad_reconciliation = reconcile_default_ads(
                ts_result,
                pc_result,
            )

            default_rules.evaluate(
                default_ad_reconciliation,
                findings_buffer,
            )

        # ----------------------------------------------------
        # Who's doing this QA2 pass -- asked once, up front, right
        # where it's used: stamps approvals below and gates the QA2
        # Sign-off checkbox further down.
        # ----------------------------------------------------

        st.subheader("QA2 Review")

        _qa2_by_cols = st.columns(2)
        record_qa2_by = _qa2_by_cols[0].text_input(
            "QA2 By", key="qa2_record_qa2_by"
        )
        record_qa2_date = _qa2_by_cols[1].date_input(
            "QA2 Date",
            value=None,
            key="qa2_record_qa2_date",
        )

        # ----------------------------------------------------
        # Review approval: REVIEW -> PASS
        #
        # A REVIEW finding is a callout, not a blocker (e.g. the TS
        # asking for a combination that's wrong on its own terms, or
        # two creatives sharing an ID with a naming mismatch). This
        # lets a QA approve one with an observation explaining why --
        # the finding's status actually becomes PASS everywhere below
        # (verdict, Findings tab, exports), with the observation
        # folded into its Reason so the record isn't lost. Rendered
        # before the verdict so an approval made this run is reflected
        # immediately, including in the same run's PDF/Excel.
        # ----------------------------------------------------

        review_findings = [
            finding for finding in findings_buffer.findings
            if finding.status.value == "REVIEW"
        ]

        if review_findings:
            with st.expander(
                f"📝 QA2 Review ({len(review_findings)})",
                expanded=True,
            ):
                st.caption(
                    "QA2 is mandatory: whoever does the second-pass "
                    "review goes through each item below, writes why "
                    "it's fine in Observation (based on the campaign "
                    "context -- the email/Wrike callouts, etc.), and "
                    "checks Approve. Only then does the finding become "
                    "PASS below, with the observation kept in the "
                    "record."
                )

                _review_base_df = pd.DataFrame(
                    [
                        {
                            "Approve": False,
                            "Rule": finding.rule_id,
                            "Placement ID": finding.placement_id,
                            "Creative ID": finding.creative_id,
                            "Finding": finding.message,
                            "Observation": "",
                        }
                        for finding in review_findings
                    ]
                )

                _edited_review_df = st.data_editor(
                    _review_base_df,
                    use_container_width=True,
                    hide_index=True,
                    disabled=[
                        "Rule", "Placement ID", "Creative ID", "Finding",
                    ],
                    key="qa2_review_approval",
                )

                review_overrides: dict[str, dict] = {}

                for finding, (_, _row) in zip(
                    review_findings, _edited_review_df.iterrows()
                ):
                    if bool(_row.get("Approve")):
                        review_overrides[finding.finding_id] = {
                            "approved": True,
                            "note": str(_row.get("Observation") or "").strip(),
                        }

                if review_overrides:
                    st.info(
                        f"{len(review_overrides)} of "
                        f"{len(review_findings)} REVIEW item(s) "
                        "approved -- counted as PASS below."
                    )

                    findings_buffer._items = apply_review_overrides(
                        findings_buffer.findings,
                        review_overrides,
                        approved_by=record_qa2_by,
                    )
                    findings_buffer._seen = {
                        finding.finding_id
                        for finding in findings_buffer._items
                    }
        else:
            st.caption(
                "📝 QA2 Review: no REVIEW items on this run, so "
                "there's nothing to approve here."
            )

        scorecard = findings_buffer.scorecard()

        # ----------------------------------------------------
        # Results header
        # ----------------------------------------------------

        show_verdict(scorecard.verdict)

        # ----------------------------------------------------
        # QA2 sign-off gate -- separate from the automated verdict.
        # Company policy: every campaign, PASSED or not, needs a
        # human QA2 approval on record before it's considered done.
        # ----------------------------------------------------

        qa2_signoff_note = st.text_area(
            "QA2 sign-off note (optional)",
            key="qa2_signoff_note",
            placeholder=(
                "e.g. Reviewed placements, tags and pixels against "
                "the TS -- approved for delivery."
            ),
        )
        qa2_signed_off = st.checkbox(
            "I reviewed this QA2 run and approve it"
            + (f" -- {record_qa2_by}" if record_qa2_by.strip() else ""),
            key="qa2_signoff_checkbox",
            disabled=not record_qa2_by.strip(),
        )

        if not record_qa2_by.strip():
            st.caption(
                "Fill in \"QA2 By\" above (QA2 Review) to enable "
                "sign-off."
            )
        elif qa2_signed_off:
            st.success(
                f"✅ Approved by {record_qa2_by} -- this run is "
                "cleared for delivery."
            )
        else:
            st.warning(
                "⏳ QA2 Sign-off pending -- every campaign needs this, "
                "even when the automated result is PASSED."
            )

        st.markdown(
            f"""
            <div class="profile-card">
                <strong>Profile used:</strong>
                {professional_profile_name(ts_result.profile)}<br>
                <strong>Detected profile:</strong>
                {professional_profile_name(detected_profile)}<br>
                <strong>Evidence:</strong>
                {detection_evidence}<br>
                <strong>Scope Guard:</strong>
                {match_result.scope_guard or "UNKNOWN"}
            </div>
            """,
            unsafe_allow_html=True,
        )

        _record_echo_fields = [
            ("Campaign", record_campaign),
            ("Request Type", record_request_type),
            ("Wrike ID", record_wrike_id),
            ("Implemented By", record_implemented_by),
            (
                "Implementation Date",
                (
                    record_implementation_date.isoformat()
                    if record_implementation_date
                    else ""
                ),
            ),
            ("QA2 By", record_qa2_by),
            (
                "QA2 Date",
                record_qa2_date.isoformat() if record_qa2_date else "",
            ),
            ("QA3 By", record_qa3_by),
            (
                "QA3 Date",
                record_qa3_date.isoformat() if record_qa3_date else "",
            ),
        ]

        _record_echo_filled = [
            (label, value)
            for label, value in _record_echo_fields
            if value
        ]

        if _record_echo_filled or record_notes:
            _record_echo_html = "<br>".join(
                f"<strong>{html.escape(label)}:</strong> "
                f"{html.escape(str(value))}"
                for label, value in _record_echo_filled
            )

            st.markdown(
                f"""
                <div class="profile-card">
                    <strong>Implementation Record</strong><br>
                    {_record_echo_html}
                </div>
                """,
                unsafe_allow_html=True,
            )

            if record_notes:
                st.markdown(
                    f"""
                    <div class="section-note">
                        <strong>Notes / Callouts:</strong><br>
                        {html.escape(record_notes).replace(chr(10), "<br>")}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

        # ----------------------------------------------------
        # Adobe Pixel & Tag Requirements
        # ----------------------------------------------------

        if (
            ts_result.profile in ("adobe_variante_a", "adobe_variante_b")
            and adobe_pixel_result is not None
        ):
            st.subheader("Pixel & Tag Requirements")

            disqo_checks = [
                check
                for check in adobe_pixel_result.checks
                if check.disqo_required
            ]

            disqo_pass = sum(
                check.result == "PASS"
                for check in disqo_checks
            )

            disqo_fail = sum(
                check.result == "FAIL"
                for check in disqo_checks
            )

            disqo_review = sum(
                check.result == "REVIEW"
                for check in disqo_checks
            )

            tag_supported = sum(
                check.result == "PASS"
                and check.tags_disqo
                for check in disqo_checks
            )

            innovid_supported = sum(
                check.result == "PASS"
                and not check.tags_disqo
                and check.innovid_disqo
                for check in disqo_checks
            )

            metric_columns = st.columns(5)

            metric_columns[0].metric(
                "DISQO required",
                len(disqo_checks),
            )

            metric_columns[1].metric(
                "Pass",
                disqo_pass,
            )

            metric_columns[2].metric(
                "Fail",
                disqo_fail,
            )

            metric_columns[3].metric(
                "Tag evidence",
                tag_supported,
            )

            metric_columns[4].metric(
                "Innovid evidence",
                innovid_supported,
            )

            if disqo_fail:
                st.error(
                    f"{disqo_fail} placements require DISQO "
                    "but no valid evidence was found."
                )
            elif disqo_review:
                st.warning(
                    f"{disqo_review} DISQO placements require review."
                )
            else:
                st.success(
                    "All DISQO requirements have valid implementation "
                    "evidence in the delivered tag files or Innovid."
                )

            disqo_rows = []

            for check in disqo_checks:
                if check.tags_disqo:
                    evidence_source = (
                        "Delivered Tag File"
                    )
                elif check.innovid_disqo:
                    evidence_source = (
                        "Innovid Active Metering / "
                        "Third Party Impression"
                    )
                else:
                    evidence_source = (
                        "No evidence found"
                    )

                disqo_rows.append(
                    {
                        "Placement ID": (
                            check.placement_id
                        ),
                        "Placement Name": (
                            check.placement_name
                        ),
                        "Site": check.site,
                        "Requirement": "DISQO",
                        "Result": check.result,
                        "Evidence Source": (
                            evidence_source
                        ),
                        "Message": check.message,
                        "Recommended Action": (
                            check.recommended_action
                            or "No correction required."
                        ),
                    }
                )

            with st.expander(
                "View DISQO placement details",
                expanded=bool(
                    disqo_fail or disqo_review
                ),
            ):
                st.dataframe(
                    pd.DataFrame(disqo_rows),
                    use_container_width=True,
                    hide_index=True,
                )

        # ----------------------------------------------------
        # DV (DoubleVerify) Tag Delivery
        # ----------------------------------------------------

        if dv_reconciliation.checks:
            st.subheader("DV (DoubleVerify) Tag Delivery")

            dv_pass = sum(
                check.result == "PASS"
                for check in dv_reconciliation.checks
            )

            dv_fail = sum(
                check.result == "FAIL"
                for check in dv_reconciliation.checks
            )

            dv_review = sum(
                check.result == "REVIEW"
                for check in dv_reconciliation.checks
            )

            dv_not_verified = sum(
                check.result == "NOT_VERIFIED"
                for check in dv_reconciliation.checks
            )

            metric_columns = st.columns(4)

            metric_columns[0].metric(
                "DV required",
                len(dv_reconciliation.checks),
            )

            metric_columns[1].metric("Pass", dv_pass)

            metric_columns[2].metric("Fail", dv_fail)

            metric_columns[3].metric(
                "Review / Not Verified",
                dv_review + dv_not_verified,
            )

            if dv_fail:
                st.error(
                    f"{dv_fail} placements require a DV tag but it "
                    "was not found or was empty."
                )
            elif dv_not_verified:
                st.warning(
                    "Upload the DV Pinnacle file to verify DV tag "
                    "delivery for these placements."
                )
            elif dv_review:
                st.warning(
                    f"{dv_review} DV placements require review."
                )
            else:
                st.success(
                    "All DV-required placements have a delivered "
                    "tag present in the Innovid tag file."
                )

            dv_rows = [
                {
                    "Placement ID": check.placement_id,
                    "Placement Name": check.placement_name,
                    "Vendors / Pixels": check.vendor_raw,
                    "Result": check.result,
                    "In DV File": (
                        "Yes" if check.in_dv_file else "No"
                    ),
                    "In Tag File": (
                        "Yes" if check.in_tag_inventory else "No"
                    ),
                    "Message": check.message,
                }
                for check in dv_reconciliation.checks
            ]

            with st.expander(
                "View DV placement details",
                expanded=bool(dv_fail or dv_review),
            ):
                st.dataframe(
                    pd.DataFrame(dv_rows),
                    use_container_width=True,
                    hide_index=True,
                )

            if dv_reconciliation.extra_dv_placements:
                st.caption(
                    "DV Pinnacle placements outside worked scope: "
                    + ", ".join(
                        dv_reconciliation.extra_dv_placements
                    )
                )

        metric_columns = st.columns(7)

        metric_columns[0].metric(
            "Worked Placements",
            match_result.expected_total,
        )

        metric_columns[1].metric(
            "Found",
            len(match_result.matched),
        )

        metric_columns[2].metric(
            "Missing",
            len(match_result.only_expected),
        )

        metric_columns[3].metric(
            "Tag Files",
            len(tags_results),
        )

        metric_columns[4].metric(
            "Errors",
            scorecard.errors,
        )

        metric_columns[5].metric(
            "Reviews",
            scorecard.reviews,
        )

        metric_columns[6].metric(
            "Not Verified",
            scorecard.not_verified,
        )

        # ----------------------------------------------------
        # Prepare per-placement data
        # ----------------------------------------------------

        matched_by_id = {
            placement_match.placement_id: placement_match
            for placement_match in match_result.matched
        }

        expected_missing_by_id = {
            expected.placement_id: expected
            for expected in match_result.only_expected
        }

        tags_by_placement = defaultdict(list)

        outside_scope_tag_rows = []

        for (
            file_name,
            tags_result,
            tag_match_result,
        ) in tag_matches:
            for tag_link in tag_match_result.links:
                tags_by_placement[
                    tag_link.placement_id
                ].append(
                    {
                        "file_name": file_name,
                        "tags_result": tags_result,
                        "link": tag_link,
                    }
                )

                if not tag_link.in_ts_scope:
                    outside_scope_tag_rows.append(
                        {
                            "File": file_name,
                            "Placement ID": tag_link.placement_id,
                            "Placement Name": (
                                tag_link.tag_row.placement_name
                            ),
                            "Dimensions": (
                                tag_link.tag_row.dimensions
                            ),
                            "Third Party ID": (
                                tag_link.tag_row.third_party_id
                            ),
                            "Status": (
                                "Outside worked scope"
                            ),
                        }
                    )

        # ----------------------------------------------------
        # Downloadable PDF report
        # ----------------------------------------------------

        placement_ids_for_pdf = sorted(
            set(matched_by_id) | set(expected_missing_by_id)
        )

        placements_rows_for_pdf = []

        for placement_id in placement_ids_for_pdf:
            placement_match = matched_by_id.get(placement_id)

            if placement_match is not None:
                expected = placement_match.expected
                actual = placement_match.actual
            else:
                expected = expected_missing_by_id[placement_id]
                actual = None

            status = placement_status(
                placement_id,
                findings_buffer,
            )

            creative_links = (
                placement_match.creative_links
                if placement_match
                else []
            )

            tag_records = tags_by_placement.get(placement_id, [])

            placements_rows_for_pdf.append(
                {
                    "Status": status,
                    "Placement ID": placement_id,
                    "Placement Name": expected.name,
                    "Request Type": expected.request_type or "-",
                    "Dimensions": expected.dims or "-",
                    "Creatives": len(creative_links),
                    "Tag Rows": len(tag_records),
                    "Found in Innovid": (
                        "Yes" if actual else "No"
                    ),
                }
            )

        evidence_images = []

        for uploaded_image in uploaded_evidence or []:
            evidence_images.append(
                (uploaded_image.name, uploaded_image.getvalue())
            )

        pdf_report_bytes = build_pdf_report(
            ReportMeta(
                verdict=scorecard.verdict,
                verdict_label=VERDICT_LABELS.get(
                    scorecard.verdict, scorecard.verdict
                ),
                profile_used=professional_profile_name(
                    ts_result.profile
                ),
                detected_profile=professional_profile_name(
                    detected_profile
                ),
                detection_evidence=detection_evidence,
                scope_guard=match_result.scope_guard or "UNKNOWN",
                scope_evidence=match_result.scope_evidence,
                ts_campaign_id=match_result.ts_campaign_id,
                export_campaign_id=match_result.export_campaign_id,
                metrics={
                    "Worked Placements": match_result.expected_total,
                    "Found": len(match_result.matched),
                    "Missing": len(match_result.only_expected),
                    "Tag Files": len(tags_results),
                    "Errors": scorecard.errors,
                    "Reviews": scorecard.reviews,
                    "Not Verified": scorecard.not_verified,
                },
                source_files=[
                    row["File"] for row in file_rows
                ],
                campaign=record_campaign,
                request_type=record_request_type,
                wrike_id=record_wrike_id,
                implemented_by=record_implemented_by,
                implementation_date=record_implementation_date,
                qa2_by=record_qa2_by,
                qa2_date=record_qa2_date,
                qa3_by=record_qa3_by,
                qa3_date=record_qa3_date,
                notes=record_notes,
                qa2_signed_off=qa2_signed_off,
                qa2_signoff_note=qa2_signoff_note,
            ),
            findings_df=findings_dataframe(
                [
                    finding
                    for finding in findings_buffer.findings
                    if finding.status.value != "PASS"
                ]
            ),
            rules_df=rule_summary_dataframe(findings_buffer),
            files_df=files_dataframe,
            placements_df=pd.DataFrame(placements_rows_for_pdf),
            evidence_images=evidence_images,
            logo_path=(
                PROJECT_ROOT
                / "ui"
                / "assets"
                / "wpp-media-logo.png.png"
            ),
        )

        tag_coverage_rows_for_excel = [
            {
                "File": file_name,
                "Campaign ID": tags_result.campaign_id,
                "Sheet": tags_result.sheet,
                "Placements": tags_result.distinct_placements,
                "Materialized Tags": tags_result.total_tags,
                "Rows In Scope": len(
                    tag_match_result.matched_to_scope
                ),
                "Rows Out of Scope": len(
                    tag_match_result.outside_scope
                ),
                "No Innovid Match": len(
                    tag_match_result.missing_in_innovid
                ),
            }
            for file_name, tags_result, tag_match_result in tag_matches
        ]

        excel_report_bytes = build_excel_report(
            ReportMeta(
                verdict=scorecard.verdict,
                verdict_label=VERDICT_LABELS.get(
                    scorecard.verdict, scorecard.verdict
                ),
                profile_used=professional_profile_name(
                    ts_result.profile
                ),
                detected_profile=professional_profile_name(
                    detected_profile
                ),
                detection_evidence=detection_evidence,
                scope_guard=match_result.scope_guard or "UNKNOWN",
                scope_evidence=match_result.scope_evidence,
                ts_campaign_id=match_result.ts_campaign_id,
                export_campaign_id=match_result.export_campaign_id,
                metrics={
                    "Worked Placements": match_result.expected_total,
                    "Found": len(match_result.matched),
                    "Missing": len(match_result.only_expected),
                    "Tag Files": len(tags_results),
                    "Errors": scorecard.errors,
                    "Reviews": scorecard.reviews,
                    "Not Verified": scorecard.not_verified,
                },
                source_files=[
                    row["File"] for row in file_rows
                ],
                campaign=record_campaign,
                request_type=record_request_type,
                wrike_id=record_wrike_id,
                implemented_by=record_implemented_by,
                implementation_date=record_implementation_date,
                qa2_by=record_qa2_by,
                qa2_date=record_qa2_date,
                qa3_by=record_qa3_by,
                qa3_date=record_qa3_date,
                notes=record_notes,
                qa2_signed_off=qa2_signed_off,
                qa2_signoff_note=qa2_signoff_note,
            ),
            findings_df=findings_dataframe(findings_buffer.findings),
            rules_df=rule_summary_dataframe(findings_buffer),
            files_df=files_dataframe,
            placements_df=pd.DataFrame(placements_rows_for_pdf),
            tag_coverage_df=(
                pd.DataFrame(tag_coverage_rows_for_excel)
                if tag_coverage_rows_for_excel else None
            ),
            logo_path=(
                PROJECT_ROOT
                / "ui"
                / "assets"
                / "wpp-media-logo.png.png"
            ),
        )

        download_columns = st.columns(2)

        download_columns[0].download_button(
            "Download PDF Report",
            data=pdf_report_bytes,
            file_name=(
                f"qa2_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            ),
            mime="application/pdf",
            use_container_width=True,
        )

        download_columns[1].download_button(
            "Download Excel Report",
            data=excel_report_bytes,
            file_name=(
                f"qa2_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
            use_container_width=True,
        )

        (
            tab_workspace,
            tab_attention,
            tab_rules,
            tab_files,
            tab_tags,
        ) = st.tabs(
            [
                "Worked Placements",
                "Findings",
                "Rules Executed",
                "Files & Extraction",
                "Tag Coverage",
            ]
        )

        # ====================================================
        # TAB: Worked Placements
        # ====================================================

        with tab_workspace:
            st.subheader(
                "Worked Placements in the Request"
            )

            st.markdown(
                """
                <div class="section-note">
                    Each row represents a placement in scope.
                    Use the arrow to expand creatives,
                    URLs, attribution, tags, and validations.
                </div>
                """,
                unsafe_allow_html=True,
            )

            filter_columns = st.columns([2, 1, 1])

            with filter_columns[0]:
                search_text = st.text_input(
                    "Search Placement ID or name",
                    placeholder=(
                        "Type an ID or part of the name"
                    ),
                )

            with filter_columns[1]:
                status_filter = st.multiselect(
                    "Result",
                    options=[
                        "FAIL",
                        "REVIEW",
                        "NOT_VERIFIED",
                        "INFO",
                        "PASS",
                    ],
                    default=[
                        "FAIL",
                        "REVIEW",
                        "NOT_VERIFIED",
                        "INFO",
                        "PASS",
                    ],
                )

            with filter_columns[2]:
                request_options = sorted(
                    {
                        (
                            item.expected.request_type
                            if item.expected
                            else ""
                        )
                        for item in match_result.matched
                    }
                    | {
                        item.request_type
                        for item in match_result.only_expected
                    }
                )

                request_options = [
                    item
                    for item in request_options
                    if item
                ]

                request_filter = st.multiselect(
                    "Request Type",
                    options=request_options,
                )

            placement_ids = sorted(
                set(matched_by_id)
                | set(expected_missing_by_id)
            )

            visible_count = 0

            for placement_id in placement_ids:
                placement_match = matched_by_id.get(
                    placement_id
                )

                if placement_match is not None:
                    expected = placement_match.expected
                    actual = placement_match.actual
                else:
                    expected = expected_missing_by_id[
                        placement_id
                    ]
                    actual = None

                status = placement_status(
                    placement_id,
                    findings_buffer,
                )

                if status not in status_filter:
                    continue

                if (
                    request_filter
                    and expected.request_type
                    not in request_filter
                ):
                    continue

                searchable_text = (
                    f"{placement_id} "
                    f"{expected.name} "
                    f"{actual.name if actual else ''}"
                ).casefold()

                if (
                    search_text
                    and search_text.casefold()
                    not in searchable_text
                ):
                    continue

                visible_count += 1

                creative_links = (
                    placement_match.creative_links
                    if placement_match
                    else []
                )

                tag_records = tags_by_placement.get(
                    placement_id,
                    [],
                )

                placement_label = (
                    f"{STATUS_ICON.get(status, '')} "
                    f"{placement_id}  |  "
                    f"{expected.request_type}  |  "
                    f"{expected.dims or '-'}  |  "
                    f"{len(creative_links)} creative(s)  |  "
                    f"{len(tag_records)} tag row(s)  |  "
                    f"{expected.name[:105]}"
                )

                with st.expander(placement_label):
                    render_status_badge(status)

                    # ----------------------------------------
                    # Placement summary header
                    # ----------------------------------------

                    placement_metrics = st.columns(5)

                    placement_metrics[0].metric(
                        "Placement ID",
                        placement_id,
                    )

                    placement_metrics[1].metric(
                        "Request",
                        expected.request_type or "-",
                    )

                    placement_metrics[2].metric(
                        "Format",
                        expected.fmt or "-",
                    )

                    placement_metrics[3].metric(
                        "Expected Creatives",
                        len(expected.creatives),
                    )

                    placement_metrics[4].metric(
                        "Tag Rows",
                        len(tag_records),
                    )

                    # ----------------------------------------
                    # Placement: TS vs Innovid
                    # ----------------------------------------

                    st.markdown(
                        "#### 1. Placement: Traffic Sheet vs Innovid"
                    )

                    # "Decision Tree" is an Adobe-specific Innovid
                    # feature (Decision_Tree_Name column). WPP Media's
                    # standard Traffic Sheets group creatives instead
                    # via the Creative Rotations sheet -- referred to
                    # here as a "Decision Set" to avoid conflating the
                    # two concepts.
                    group_label = (
                        "Decision Tree"
                        if ts_result.profile == "adobe_variante_a"
                        else "Decision Set (Creative Rotation)"
                    )

                    placement_comparisons = [
                        comparison_row(
                            "Placement ID",
                            expected.placement_id,
                            (
                                actual.placement_id
                                if actual
                                else ""
                            ),
                        ),
                        comparison_row(
                            "Placement Name",
                            expected.name,
                            (
                                (
                                    actual.name_norm
                                    or actual.name
                                )
                                if actual
                                else ""
                            ),
                        ),
                        comparison_row(
                            "Site",
                            expected.site,
                            actual.site if actual else "",
                            fuzzy=True,
                        ),
                        comparison_row(
                            "Dimensions",
                            expected.dims,
                            actual.dims if actual else "",
                            normalizer=norm_dims,
                        ),
                        comparison_row(
                            "Start Date",
                            expected.start,
                            actual.start if actual else "",
                        ),
                        comparison_row(
                            "End Date",
                            expected.end,
                            actual.end if actual else "",
                        ),
                        comparison_row(
                            group_label,
                            expected.group_name,
                            (
                                actual.group_name
                                if actual
                                else ""
                            ),
                            optional=(
                                not bool(expected.group_name)
                                # 1x1s are direct-assigned in Innovid.
                                # The TS "Decision Set" name for a 1x1
                                # is a reference only -- it is never a
                                # real Innovid Decision Tree and never
                                # shows up in the export, so it isn't
                                # something to verify.
                                or norm_dims(expected.dims) == "1x1"
                            ),
                        ),
                    ]

                    st.dataframe(
                        pd.DataFrame(
                            placement_comparisons
                        ),
                        use_container_width=True,
                        hide_index=True,
                    )

                    # ----------------------------------------
                    # Creatives
                    # ----------------------------------------

                    st.markdown(
                        "#### 2. Creatives & Assignment"
                    )

                    creative_rows = []

                    if placement_match is not None:
                        for creative_link in creative_links:
                            expected_creative = (
                                creative_link.expected
                            )

                            actual_creative = (
                                creative_link.actual
                            )

                            creative_rows.append(
                                {
                                    "Intent": (
                                        expected_creative.intent
                                    ),
                                    "Expected Creative": (
                                        expected_creative.name
                                    ),
                                    "Found Creative": (
                                        (
                                            actual_creative.filename
                                            or actual_creative.name
                                        )
                                        if actual_creative
                                        else ""
                                    ),
                                    "Creative ID": (
                                        actual_creative.creative_id
                                        if actual_creative
                                        else (
                                            expected_creative.creative_id
                                        )
                                    ),
                                    "Match key": (
                                        creative_link.trace.winner
                                        or "-"
                                    ),
                                    "Confidence": (
                                        creative_link.confidence
                                    ),
                                    "Status": (
                                        actual_creative.state_label
                                        if actual_creative
                                        # Un creativo en rojo es una
                                        # desasignacion: que no este en
                                        # el export es la confirmacion
                                        # de que se hizo, no una falta.
                                        else (
                                            "Removed (confirmed)"
                                            if expected_creative.intent
                                            == RED
                                            else "MISSING"
                                        )
                                    ),
                                    # En un creativo que se removio no hay
                                    # URL ni atribucion que revisar: se
                                    # fue. Marcarlo NOT_VERIFIED hacia
                                    # parecer pendiente algo que ya esta
                                    # resuelto.
                                    "URL": (
                                        "N/A (removed)"
                                        if expected_creative.intent == RED
                                        else (
                                            creative_link.url.result
                                            if creative_link.url
                                            else "NOT_VERIFIED"
                                        )
                                    ),
                                    "Attribution": (
                                        "N/A (removed)"
                                        if expected_creative.intent == RED
                                        else (
                                            creative_link.triangle.result
                                            if creative_link.triangle
                                            else "NOT_VERIFIED"
                                        )
                                    ),
                                }
                            )

                    if creative_rows:
                        st.dataframe(
                            pd.DataFrame(creative_rows),
                            use_container_width=True,
                            hide_index=True,
                        )
                    else:
                        st.info(
                            "No individual creatives are "
                            "expected for this placement."
                        )

                    # ----------------------------------------
                    # URL and attribution
                    # ----------------------------------------

                    st.markdown(
                        "#### 3. URLs & Attribution"
                    )

                    url_links = [
                        creative_link
                        for creative_link in creative_links
                        if (
                            creative_link.url is not None
                            or creative_link.triangle is not None
                        )
                    ]

                    if not url_links:
                        st.info(
                            "No URL or attribution validations "
                            "were run for this placement."
                        )

                    for creative_index, creative_link in enumerate(
                        url_links,
                        start=1,
                    ):
                        creative_title = (
                            creative_link.expected.name
                            or (
                                creative_link.actual.filename
                                if creative_link.actual
                                else ""
                            )
                            or f"Creative {creative_index}"
                        )

                        url_result = (
                            creative_link.url.result
                            if creative_link.url
                            else "NOT_VERIFIED"
                        )

                        triangle_result = (
                            creative_link.triangle.result
                            if creative_link.triangle
                            else "NOT_VERIFIED"
                        )

                        with st.expander(
                            f"{STATUS_ICON.get('PASS' if url_result == 'MATCH' else 'REVIEW', '')} "
                            f"{creative_title[:120]}  |  "
                            f"URL: {url_result}  |  "
                            f"Attribution: {triangle_result}"
                        ):
                            if creative_link.url is not None:
                                st.write(
                                    "**URL Result:**",
                                    creative_link.url.result,
                                )

                                url_columns = st.columns(2)

                                with url_columns[0]:
                                    st.caption(
                                        "URL expected in Traffic Sheet"
                                    )
                                    st.code(
                                        creative_link.url.expected.raw
                                        or "Pending confirmation",
                                        language=None,
                                    )

                                with url_columns[1]:
                                    st.caption(
                                        "URL found in Innovid"
                                    )
                                    st.code(
                                        creative_link.url.actual.raw
                                        or "Pending confirmation",
                                        language=None,
                                    )

                                if creative_link.url.note:
                                    st.caption(
                                        creative_link.url.note
                                    )

                            if creative_link.triangle is not None:
                                triangle = (
                                    creative_link.triangle
                                )

                                st.write(
                                    "**Attribution Triangle:**",
                                    triangle.result,
                                )

                                triangle_columns = st.columns(3)

                                triangle_columns[0].metric(
                                    "CGEN in TS",
                                    triangle.ts or "-",
                                )

                                triangle_columns[1].metric(
                                    "Third Party ID",
                                    triangle.export or "-",
                                )

                                triangle_columns[2].metric(
                                    "sdid in URL",
                                    triangle.url or "-",
                                )

                                st.caption(
                                    triangle.note
                                    or "No additional detail."
                                )

                    # ----------------------------------------
                    # Extras
                    # ----------------------------------------

                    if (
                        placement_match is not None
                        and placement_match.actual_extra
                    ):
                        st.markdown(
                            "#### 4. Extra Creatives in Innovid"
                        )

                        extra_rows = []

                        for extra in (
                            placement_match.actual_extra
                        ):
                            extra_rows.append(
                                {
                                    "Creative ID": (
                                        extra.creative_id
                                    ),
                                    "Name or filename": (
                                        extra.filename
                                        or extra.name
                                    ),
                                    "Status": extra.state_label,
                                    "Running": (
                                        "Yes"
                                        if extra.running
                                        else "No"
                                    ),
                                    group_label: (
                                        extra.group_name
                                    ),
                                    "Third Party ID": (
                                        extra.third_party_id
                                    ),
                                }
                            )

                        st.dataframe(
                            pd.DataFrame(extra_rows),
                            use_container_width=True,
                            hide_index=True,
                        )

                    # ----------------------------------------
                    # Delivered tags
                    # ----------------------------------------

                    st.markdown(
                        "#### 5. Tag Files"
                    )

                    if not tag_records:
                        st.info(
                            "No Tag file containing this "
                            "Placement ID was uploaded."
                        )
                    else:
                        tag_summary_rows = []

                        for tag_record in tag_records:
                            tag_row = (
                                tag_record["link"].tag_row
                            )

                            tag_summary_rows.append(
                                {
                                    "File": (
                                        tag_record["file_name"]
                                    ),
                                    "Placement ID": (
                                        tag_row.placement_id
                                    ),
                                    "Placement Name": (
                                        tag_row.placement_name
                                    ),
                                    "Dimensions": (
                                        tag_row.dimensions
                                    ),
                                    "Third Party ID": (
                                        tag_row.third_party_id
                                    ),
                                    "Prisma ID": (
                                        tag_row.prisma_id
                                    ),
                                    "Tag Count": (
                                        tag_row.tag_count
                                    ),
                                    "Detected Types": ", ".join(
                                        sorted(
                                            {
                                                tag.tag_type
                                                for tag
                                                in tag_row.tags
                                            }
                                        )
                                    ),
                                }
                            )

                        st.dataframe(
                            pd.DataFrame(
                                tag_summary_rows
                            ),
                            use_container_width=True,
                            hide_index=True,
                        )

                        for tag_record in tag_records:
                            file_name = (
                                tag_record["file_name"]
                            )

                            tag_row = (
                                tag_record["link"].tag_row
                            )

                            with st.expander(
                                f"View Full Tags | {file_name}"
                            ):
                                tag_metadata_columns = (
                                    st.columns(4)
                                )

                                tag_metadata_columns[0].metric(
                                    "Placement ID",
                                    tag_row.placement_id,
                                )

                                tag_metadata_columns[1].metric(
                                    "Dimensions",
                                    tag_row.dimensions or "-",
                                )

                                tag_metadata_columns[2].metric(
                                    "Third Party ID",
                                    (
                                        tag_row.third_party_id
                                        or "-"
                                    ),
                                )

                                tag_metadata_columns[3].metric(
                                    "Prisma ID",
                                    tag_row.prisma_id or "-",
                                )

                                for tag in tag_row.tags:
                                    st.markdown(
                                        f"##### {tag.column_name} "
                                        f"| {tag.tag_type}"
                                    )

                                    tag_information = pd.DataFrame(
                                        [
                                            {
                                                "Field": (
                                                    "Campaign IDs"
                                                ),
                                                "Value": ", ".join(
                                                    tag.campaign_ids
                                                ),
                                            },
                                            {
                                                "Field": (
                                                    "Placement IDs"
                                                ),
                                                "Value": ", ".join(
                                                    tag.placement_ids
                                                ),
                                            },
                                            {
                                                "Field": "Hosts",
                                                "Value": ", ".join(
                                                    tag.hosts
                                                ),
                                            },
                                            {
                                                "Field": "Macros",
                                                "Value": ", ".join(
                                                    tag.macros
                                                ),
                                            },
                                            {
                                                "Field": "Widths",
                                                "Value": ", ".join(
                                                    tag.widths
                                                ),
                                            },
                                            {
                                                "Field": "Heights",
                                                "Value": ", ".join(
                                                    tag.heights
                                                ),
                                            },
                                        ]
                                    )

                                    st.dataframe(
                                        tag_information,
                                        use_container_width=True,
                                        hide_index=True,
                                    )

                                    st.code(
                                        tag.raw,
                                        language="html",
                                    )

                    # ----------------------------------------
                    # Placement findings
                    # ----------------------------------------

                    st.markdown(
                        "#### 6. Validations Run"
                    )

                    current_findings = placement_findings(
                        placement_id,
                        findings_buffer,
                    )

                    current_findings_df = (
                        findings_dataframe(
                            current_findings
                        )
                    )

                    if current_findings_df.empty:
                        st.info(
                            "No rule results are associated "
                            "with this placement."
                        )
                    else:
                        st.dataframe(
                            current_findings_df,
                            use_container_width=True,
                            hide_index=True,
                            height=min(
                                480,
                                80
                                + (
                                    len(
                                        current_findings_df
                                    )
                                    * 35
                                ),
                            ),
                        )

            if visible_count == 0:
                st.warning(
                    "No placements match the "
                    "selected filters."
                )
            else:
                st.caption(
                    f"Visible placements: {visible_count}"
                )

        # ====================================================
        # TAB: Findings
        # ====================================================

        with tab_attention:
            st.subheader(
                "Findings That Require Attention"
            )

            attention_findings = [
                finding
                for finding in findings_buffer.findings
                if finding.status.value != "PASS"
            ]

            attention_df = findings_dataframe(attention_findings)

            if attention_df.empty:
                st.success(
                    "No errors, reviews, or pending "
                    "validations were detected."
                )
            else:
                selected_attention_statuses = (
                    st.multiselect(
                        "Filter status",
                        options=[
                            "FAIL",
                            "REVIEW",
                            "NOT_VERIFIED",
                            "INFO",
                        ],
                        default=[
                            "FAIL",
                            "REVIEW",
                            "NOT_VERIFIED",
                            "INFO",
                        ],
                    )
                )

                filtered_attention_df = (
                    attention_df[
                        attention_df["Status"].isin(
                            selected_attention_statuses
                        )
                    ]
                )

                # Un mismo problema repetido en 60 placements son 60
                # filas identicas en el detalle. Agrupado se lee de una:
                # cuantos placements comparten cada hallazgo y cuales.
                grouped_rows = []

                for (rule, status, message), group in (
                    filtered_attention_df.groupby(
                        ["Rule", "Status", "Message"],
                        sort=False,
                    )
                ):
                    placements_in_group = [
                        str(value)
                        for value in group["Placement ID"].tolist()
                        if str(value).strip()
                    ]

                    grouped_rows.append(
                        {
                            "Placements": len(group),
                            "Status": status,
                            "Rule": rule,
                            "Finding": message,
                            "Placement IDs": ", ".join(
                                dict.fromkeys(placements_in_group)
                            ),
                        }
                    )

                grouped_df = pd.DataFrame(grouped_rows)

                if not grouped_df.empty:
                    grouped_df = grouped_df.sort_values(
                        "Placements",
                        ascending=False,
                        kind="stable",
                    )

                    st.markdown("#### Grouped by finding")
                    st.caption(
                        f"{len(grouped_df)} distinct findings across "
                        f"{len(filtered_attention_df)} rows."
                    )

                    st.dataframe(
                        grouped_df,
                        use_container_width=True,
                        hide_index=True,
                    )

                    st.download_button(
                        "Download Grouped Findings CSV",
                        data=grouped_df.to_csv(
                            index=False,
                            encoding="utf-8-sig",
                        ),
                        file_name="qa2_findings_grouped.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )

                st.markdown("#### Every finding")

                st.dataframe(
                    filtered_attention_df,
                    use_container_width=True,
                    hide_index=True,
                    height=620,
                )

                st.download_button(
                    "Download Findings CSV",
                    data=filtered_attention_df.to_csv(
                        index=False,
                        encoding="utf-8-sig",
                    ),
                    file_name="qa2_findings.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

        # ====================================================
        # TAB: Rules Executed
        # ====================================================

        with tab_rules:
            st.subheader(
                "Rule Execution Coverage"
            )

            rules_dataframe = (
                rule_summary_dataframe(
                    findings_buffer
                )
            )

            if rules_dataframe.empty:
                st.warning(
                    "The Rule Engine did not emit any results."
                )
            else:
                st.dataframe(
                    rules_dataframe,
                    use_container_width=True,
                    hide_index=True,
                )

        # ====================================================
        # TAB: Files & Extraction
        # ====================================================

        with tab_files:
            st.subheader(
                "Document Understanding"
            )

            st.dataframe(
                files_dataframe,
                use_container_width=True,
                hide_index=True,
            )

            st.markdown("#### Traffic Sheet")

            st.write(
                f"**Profile used:** "
                f"{professional_profile_name(ts_result.profile)}"
            )

            st.write(
                f"**Detected profile:** "
                f"{professional_profile_name(detected_profile)}"
            )

            st.write(
                f"**Evidence:** {detection_evidence}"
            )

            st.write(
                f"**Worked placements:** "
                f"{len(ts_result.worked)}"
            )

            st.markdown("#### Scope Guard")

            scope_columns = st.columns(3)

            scope_columns[0].metric(
                "Result",
                match_result.scope_guard or "UNKNOWN",
            )

            scope_columns[1].metric(
                "Campaign ID TS",
                match_result.ts_campaign_id or "-",
            )

            scope_columns[2].metric(
                "Campaign ID Innovid",
                match_result.export_campaign_id or "-",
            )

            st.caption(
                match_result.scope_evidence
            )

            st.markdown(
                "#### Extraction Anomalies"
            )

            if all_anomaly_rows:
                st.dataframe(
                    pd.DataFrame(
                        all_anomaly_rows
                    ),
                    use_container_width=True,
                    hide_index=True,
                )
            else:
                st.success(
                    "No extraction anomalies "
                    "were detected."
                )

        # ====================================================
        # TAB: Tag Coverage
        # ====================================================

        with tab_tags:
            st.subheader(
                "Tag File Coverage"
            )

            if not tag_matches:
                st.info(
                    "No Tag files were uploaded."
                )
            else:
                tag_file_summary = []

                for (
                    file_name,
                    tags_result,
                    tag_match_result,
                ) in tag_matches:
                    tag_file_summary.append(
                        {
                            "File": file_name,
                            "Campaign ID": (
                                tags_result.campaign_id
                            ),
                            "Sheet": tags_result.sheet,
                            "Placements": (
                                tags_result.distinct_placements
                            ),
                            "Materialized Tags": (
                                tags_result.total_tags
                            ),
                            "Rows In Scope": (
                                len(
                                    tag_match_result
                                    .matched_to_scope
                                )
                            ),
                            "Rows Out of Scope": (
                                len(
                                    tag_match_result
                                    .outside_scope
                                )
                            ),
                            "No Innovid Match": (
                                len(
                                    tag_match_result
                                    .missing_in_innovid
                                )
                            ),
                        }
                    )

                st.dataframe(
                    pd.DataFrame(
                        tag_file_summary
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                st.markdown(
                    "#### Tag Placements "
                    "Outside Worked Scope"
                )

                if outside_scope_tag_rows:
                    st.dataframe(
                        pd.DataFrame(
                            outside_scope_tag_rows
                        ),
                        use_container_width=True,
                        hide_index=True,
                        height=500,
                    )
                else:
                    st.success(
                        "All placements found in Tags "
                        "belong to the worked scope."
                    )

    except Exception as error:
        st.error(
            "QA2 could not complete processing."
        )

        st.exception(error)
