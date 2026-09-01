"""
Parser for DoubleVerify (DV) Pinnacle site-served tag exports.

These are delivered as a separate file from the Innovid tag export
(ft_tags) whenever a placement's Traffic Sheet "Vendors / Pixels"
value mentions DV. One row per placement, with the DV tag markup in
a Display or Video column depending on format.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from core.extraction import (
    FieldSpec,
    SheetSpec,
    find_header_row,
    map_columns,
    read_sheet,
    resolve_sheet,
)
from core.normalize import clean_id
from core.provenance import Anomaly

DV_FIELDS = [
    FieldSpec("placement_id", ["Placement ID"], required=True, kind="id"),
    FieldSpec("placement_name", ["Placement Name"]),
    FieldSpec(
        "display_tag",
        ["Display Site-Served Tags", "Display Site Served Tags"],
    ),
    FieldSpec(
        "video_tag",
        ["Video Site-Served Tags", "Video Site Served Tags"],
    ),
]

DV_SPEC = SheetSpec(
    sheet_aliases=["Site-Served Tags", "Site Served Tags", "DV Tags"],
    header_mode="signature",
    signature_scan_rows=40,
    signature_min_matches=3,
    fields=DV_FIELDS,
)


@dataclass
class DVTagRow:
    row: int
    placement_id: str = ""
    placement_name: str = ""
    display_tag: str = ""
    video_tag: str = ""

    @property
    def has_tag(self) -> bool:
        return bool(self.display_tag.strip() or self.video_tag.strip())


@dataclass
class DVTagsResult:
    path: str = ""
    sheet: str = ""
    rows: list[DVTagRow] = field(default_factory=list)
    anomalies: list[Anomaly] = field(default_factory=list)

    @property
    def fatal(self) -> bool:
        return any(a.severity == "FATAL" for a in self.anomalies)

    @property
    def placement_ids(self) -> set[str]:
        return {r.placement_id for r in self.rows if r.placement_id}


def parse_dv_tags(path: Path) -> DVTagsResult:
    result = DVTagsResult(path=str(path))

    wb = load_workbook(path, data_only=True, keep_vba=False, read_only=True)
    sheet_name = resolve_sheet(wb.sheetnames, DV_SPEC.sheet_aliases)
    wb.close()

    if sheet_name is None:
        result.anomalies.append(
            Anomaly(
                "DV-SHEET-NOT-FOUND",
                "FATAL",
                "No sheet matching a DV Pinnacle site-served tags "
                "export was found (expected 'Site-Served Tags').",
            )
        )
        return result

    result.sheet = sheet_name
    grid, grid_anomalies = read_sheet(path, sheet_name)
    result.anomalies.extend(grid_anomalies)
    if result.fatal:
        return result

    header_result, header_anomalies = find_header_row(grid, DV_SPEC)
    result.anomalies.extend(header_anomalies)
    if header_result.row is None:
        return result

    cmap, column_anomalies = map_columns(grid, header_result.row, DV_SPEC)
    result.anomalies.extend(column_anomalies)
    if result.fatal:
        return result

    seen_placements: set[str] = set()

    for row_num in range(header_result.row + 1, grid.max_row + 1):
        cells = grid.row_cells(row_num)
        if not cells:
            continue

        pid_col = cmap.col("placement_id")
        placement_id = clean_id(
            cells[pid_col].raw if pid_col in cells else None
        )
        if not placement_id:
            continue

        # The template repeats the header row once; skip a duplicate.
        if placement_id in ("Placement ID",):
            continue

        name_col = cmap.col("placement_name")
        display_col = cmap.col("display_tag")
        video_col = cmap.col("video_tag")

        row = DVTagRow(
            row=row_num,
            placement_id=placement_id,
            placement_name=(
                cells[name_col].text if name_col in cells else ""
            ),
            display_tag=(
                cells[display_col].text if display_col in cells else ""
            ),
            video_tag=(
                cells[video_col].text if video_col in cells else ""
            ),
        )
        result.rows.append(row)
        seen_placements.add(placement_id)

    if not result.rows:
        result.anomalies.append(
            Anomaly(
                "DV-NO-DATA",
                "FATAL",
                "The file contains no rows with Placement ID.",
            )
        )

    return result
