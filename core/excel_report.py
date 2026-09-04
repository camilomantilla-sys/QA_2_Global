"""
Branded multi-tab Excel export of a QA2 run.

Meant as the detailed working record analysts filter/pivot on day to
day; the PDF report is the print-and-sign record for QA3. Both are
built from the same underlying dataframes so they never disagree.
"""
from __future__ import annotations

import io
from pathlib import Path

import pandas as pd  # type: ignore
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.pdf_report import ReportMeta

WPP_INDIGO = "4B5EEA"
WPP_INDIGO_DARK = "2C36A8"
WPP_INK = "171B2E"
WPP_MUTED = "64748B"
WPP_BG = "F5F7FD"

STATUS_FILLS = {
    "PASS": "D9F7EC",
    "FAIL": "FDE2E2",
    "REVIEW": "FDECD2",
    "NOT_VERIFIED": "E9EBF3",
    "INFO": "D9F2FA",
}

STATUS_FONT_COLORS = {
    "PASS": "0FA97C",
    "FAIL": "DC2626",
    "REVIEW": "B45309",
    "NOT_VERIFIED": "475569",
    "INFO": "0E7490",
}

HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
HEADER_FILL = PatternFill(
    start_color=WPP_INDIGO, end_color=WPP_INDIGO, fill_type="solid"
)
TITLE_FONT = Font(color=WPP_INDIGO_DARK, bold=True, size=15)
LABEL_FONT = Font(color=WPP_MUTED, bold=True, size=9)
VALUE_FONT = Font(color=WPP_INK, size=10)
BODY_FONT = Font(color=WPP_INK, size=9.5)


def _write_table(
    ws: Worksheet, df: pd.DataFrame, start_row: int = 1,
    status_col: str | None = None,
) -> int:
    """Write a dataframe as a styled, filterable table. Returns next free row."""
    if df.empty:
        ws.cell(
            row=start_row, column=1, value="No rows."
        ).font = BODY_FONT
        return start_row + 2

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    status_idx = (
        list(df.columns).index(status_col) + 1
        if status_col and status_col in df.columns
        else None
    )

    for row_offset, (_, row) in enumerate(df.iterrows(), start=1):
        row_num = start_row + row_offset
        for col_idx, col_name in enumerate(df.columns, start=1):
            value = row[col_name]
            if pd.isna(value):
                value = ""
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(
                vertical="top", wrap_text=(col_idx != status_idx)
            )
            if (
                status_idx
                and col_idx == status_idx
                and str(value) in STATUS_FILLS
            ):
                cell.fill = PatternFill(
                    start_color=STATUS_FILLS[str(value)],
                    end_color=STATUS_FILLS[str(value)],
                    fill_type="solid",
                )
                cell.font = Font(
                    color=STATUS_FONT_COLORS[str(value)],
                    bold=True,
                    size=9.5,
                )

    last_row = start_row + len(df)
    last_col = len(df.columns)
    ws.auto_filter.ref = (
        f"A{start_row}:{get_column_letter(last_col)}{last_row}"
    )
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col_name))]
            + [
                len(str(v)) for v in df[col_name].head(300).tolist()
            ]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 2, 10), 60
        )

    return last_row + 2


def _blank(value, fallback: str = "") -> str:
    return str(value) if value else fallback


def _summary_sheet(wb: Workbook, meta: ReportMeta, logo_path: Path | None):
    ws = wb.active
    ws.title = "Summary"

    row = 1
    if logo_path is not None and logo_path.exists():
        img = XLImage(str(logo_path))
        img.width = 180
        img.height = 100
        ws.add_image(img, "A1")
        row = 7

    ws.cell(row=row, column=1, value="INNOVID QA2 AUTOMATION").font = (
        TITLE_FONT
    )
    row += 1
    ws.cell(
        row=row, column=1,
        value=f"Overall result: {meta.verdict_label}",
    ).font = Font(
        color=STATUS_FONT_COLORS.get(
            "FAIL"
            if meta.verdict in ("FAILED", "BLOCKED")
            else "REVIEW" if meta.verdict == "NEEDS_REVIEW" else "PASS",
            WPP_INK,
        ),
        bold=True, size=12,
    )
    row += 1

    signoff_label = (
        f"QA2 Sign-off: Approved by {meta.qa2_by or 'QA2'}"
        + (f" on {meta.qa2_date.isoformat()}" if meta.qa2_date else "")
        if meta.qa2_signed_off
        else "QA2 Sign-off: PENDING"
    )
    signoff_cell = ws.cell(row=row, column=1, value=signoff_label)
    signoff_cell.font = Font(
        color=STATUS_FONT_COLORS.get(
            "PASS" if meta.qa2_signed_off else "REVIEW", WPP_INK
        ),
        bold=True, size=11,
    )
    signoff_cell.fill = PatternFill(
        "solid",
        fgColor=STATUS_FILLS["PASS" if meta.qa2_signed_off else "REVIEW"],
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    if not meta.qa2_signed_off:
        ws.cell(
            row=row, column=1,
            value=(
                "This is the record of QA2 approval for this campaign "
                "-- to approve, edit the cell above directly here in "
                "SharePoint/Excel Web, e.g. \"QA2 Sign-off: Approved "
                "by Camilo Mantilla on 2026-09-05\"."
            ),
        ).font = Font(color=WPP_MUTED, italic=True, size=9)
        row += 1

    if meta.qa2_signoff_note:
        ws.cell(row=row, column=1, value=meta.qa2_signoff_note).font = (
            BODY_FONT
        )
        row += 1
    row += 1

    info_rows = [
        ("Profile used", meta.profile_used),
        ("Detected profile", meta.detected_profile),
        ("Evidence", meta.detection_evidence),
        ("Scope Guard", f"{meta.scope_guard} — {meta.scope_evidence}"),
        (
            "Campaign ID (TS / Innovid)",
            f"{meta.ts_campaign_id or '-'} / "
            f"{meta.export_campaign_id or '-'}",
        ),
        ("Generated", meta.generated_at.strftime("%Y-%m-%d %H:%M")),
        ("Source files", ", ".join(meta.source_files)),
    ]
    for label, value in info_rows:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = VALUE_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Metrics").font = Font(
        color=WPP_INDIGO_DARK, bold=True, size=11
    )
    row += 1
    for label, value in meta.metrics.items():
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = Font(
            color=WPP_INDIGO_DARK, bold=True, size=11
        )
        row += 1

    row += 1
    ws.cell(
        row=row, column=1, value="Implementation Record"
    ).font = Font(color=WPP_INDIGO_DARK, bold=True, size=11)
    row += 1
    record_rows = [
        ("Campaign", meta.campaign),
        ("Request Type", meta.request_type),
        ("Wrike ID", meta.wrike_id),
        ("Implemented By", meta.implemented_by),
        (
            "Implementation Date",
            meta.implementation_date.isoformat()
            if meta.implementation_date else "",
        ),
        ("QA2 By", meta.qa2_by),
        ("QA2 Date", meta.qa2_date.isoformat() if meta.qa2_date else ""),
        ("QA3 By", meta.qa3_by),
        ("QA3 Date", meta.qa3_date.isoformat() if meta.qa3_date else ""),
    ]
    for label, value in record_rows:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=_blank(value, "—")).font = (
            VALUE_FONT
        )
        row += 1

    if meta.notes:
        row += 1
        ws.cell(row=row, column=1, value="Notes / Callouts").font = Font(
            color=WPP_INDIGO_DARK, bold=True, size=11
        )
        row += 1
        cell = ws.cell(row=row, column=1, value=meta.notes)
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row + 3, end_column=6
        )

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60


def build_excel_report(
    meta: ReportMeta,
    findings_df: pd.DataFrame,
    rules_df: pd.DataFrame,
    files_df: pd.DataFrame,
    placements_df: pd.DataFrame | None = None,
    tag_coverage_df: pd.DataFrame | None = None,
    logo_path: Path | None = None,
) -> bytes:
    """Render the full branded QA2 workbook and return it as XLSX bytes."""
    wb = Workbook()

    _summary_sheet(wb, meta, logo_path)

    ws = wb.create_sheet("Worked Placements")
    _write_table(
        ws, placements_df if placements_df is not None else pd.DataFrame(),
        status_col="Status",
    )

    ws = wb.create_sheet("Findings")
    _write_table(ws, findings_df, status_col="Status")

    ws = wb.create_sheet("Rules Executed")
    _write_table(ws, rules_df)

    ws = wb.create_sheet("Files & Extraction")
    _write_table(ws, files_df, status_col="Status")

    if tag_coverage_df is not None:
        ws = wb.create_sheet("Tag Coverage")
        _write_table(ws, tag_coverage_df)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
