"""
Capa de EXTRACCION (L1). Aqui estaba el bloqueo original.

Cinco componentes:
  1. SheetReader        - lee celdas con provenance (fill, hyperlink, tipo)
  2. Header Anchor      - fixed_row | signature | labels
  3. Column Mapper      - por ALIAS, jamas por posicion  (mata B1 y B2)
  4. Block Segmenter    - clasifica filas antes de convertirlas en dato
  5. Merged-cell        - fill-down explicito y auditado
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.provenance import EXCEL_ERRORS, Anomaly, Cell, CellRef
from core.normalize import norm_key, norm_text

# ---------------------------------------------------------------- tipos de fila

ROW_DATA = "DATA"
ROW_BANNER = "BANNER"
ROW_TEMPLATE = "TEMPLATE"
ROW_SEPARATOR = "SEPARATOR"
ROW_CORRUPT = "CORRUPT"

CORRUPT_MARKERS = ("[object object]",)

# ---------------------------------------------------------------- contrato de campo

@dataclass
class FieldSpec:
    """Contrato de un campo canonico. La unica fuente de verdad del mapeo."""
    name: str
    aliases: list[str]
    required: bool = False
    kind: str = "text"       # text | id | date | dims | url | bool | number
    multi: bool = False      # familias tipo Clicktag_1..20

    def keys(self) -> set[str]:
        return {norm_key(a) for a in ([self.name] + self.aliases)}

@dataclass
class SheetSpec:
    """Como se lee una hoja concreta."""
    sheet_aliases: list[str]
    header_mode: str = "signature"   # fixed_row | signature | labels
    header_row: int | None = None
    signature_scan_rows: int = 30
    signature_min_matches: int = 4
    fields: list[FieldSpec] = field(default_factory=list)
    fill_down: list[str] = field(default_factory=list)
    ignore_row_patterns: list[str] = field(default_factory=list)
    metadata_labels: list[str] = field(default_factory=list)
    metadata_scan_rows: int = 9
    entity_identity_any: list[str] = field(default_factory=list)
    entity_identity_when: list[str] = field(default_factory=list)   # <-- NUEVA

# ---------------------------------------------------------------- lectura cruda

@dataclass
class SheetGrid:
    """La hoja completa, celda por celda, con provenance."""
    doc: str
    sheet: str
    rows: dict[int, dict[int, Cell]] = field(default_factory=dict)
    max_row: int = 0
    max_col: int = 0
    merged_applied: int = 0

    def cell(self, row: int, col: int) -> Cell | None:
        return self.rows.get(row, {}).get(col)

    def row_cells(self, row: int) -> dict[int, Cell]:
        return self.rows.get(row, {})

    def row_text(self, row: int) -> list[str]:
        return [c.text for c in self.rows.get(row, {}).values() if c.text]

def _fill_rgb(cell_obj: Any) -> str | None:
    """
    Extrae el relleno solido como token.
    Formatos: 'AARRGGBB' | 'IDX:n' | 'THEME:n:tint'

    Los getattr se validan con isinstance porque openpyxl devuelve el texto
    del descriptor cuando el atributo no esta definido (bug B31).
    """
    try:
        fill = cell_obj.fill
        if fill is None or fill.patternType != "solid":
            return None
        fg = fill.fgColor

        rgb = getattr(fg, "rgb", None)
        if isinstance(rgb, str) and len(rgb) in (6, 8):
            body = rgb[2:] if len(rgb) == 8 else rgb
            if all(c in "0123456789abcdefABCDEF" for c in body):
                return rgb.upper()

        idx = getattr(fg, "indexed", None)
        if isinstance(idx, int):
            return f"IDX:{idx}"

        th = getattr(fg, "theme", None)
        if isinstance(th, int):
            tint = getattr(fg, "tint", 0.0)
            if not isinstance(tint, (int, float)):
                tint = 0.0
            return f"THEME:{th}:{round(float(tint), 4)}"

        return None
    except Exception:
        return None

def read_sheet(path: Path, sheet_name: str,
               capture_fill: bool = True,
               max_rows_cap: int = 200_000) -> tuple[SheetGrid, list[Anomaly]]:
    """
    Lee una hoja completa. Soporta .xlsx y .xlsm.

    max_rows_cap: proteccion contra ws.max_row inflado. Excel reporta
    1.048.576 cuando hay formato aplicado a columnas completas, y sin este
    tope el bucle es inviable.
    """
    anomalies: list[Anomaly] = []
    doc = path.name

    wb = load_workbook(path, data_only=True, keep_vba=False)
    if sheet_name not in wb.sheetnames:
        anomalies.append(Anomaly("EXT-SHEET-MISSING", "FATAL",
                                 f"Hoja '{sheet_name}' no existe. "
                                 f"Disponibles: {wb.sheetnames}"))
        return SheetGrid(doc=doc, sheet=sheet_name), anomalies

    ws = wb[sheet_name]
    declared_rows = ws.max_row or 0
    declared_cols = ws.max_column or 0

    read_limit = min(declared_rows, max_rows_cap)
    if declared_rows > max_rows_cap:
        anomalies.append(Anomaly(
            "EXT-MAXROW-INFLATED", "WARNING",
            f"La hoja declara {declared_rows:,} filas (formato aplicado a "
            f"columnas completas). Se leen las primeras {max_rows_cap:,}.",
            detail={"declared": declared_rows, "read": max_rows_cap},
        ))

    grid = SheetGrid(doc=doc, sheet=sheet_name,
                     max_row=0, max_col=declared_cols)

    last_content_row = 0

    for row_obj in ws.iter_rows(min_row=1, max_row=read_limit):
        row_has_content = False
        for c in row_obj:
            fill = _fill_rgb(c) if capture_fill else None
            if c.value is None and fill is None:
                continue
            if c.value is not None:
                row_has_content = True
            text = norm_text(str(c.value)) if c.value is not None else ""
            is_err = text in EXCEL_ERRORS
            cell = Cell(
                ref=CellRef(doc, sheet_name, c.row, c.column,
                            get_column_letter(c.column)),
                raw=c.value,
                text=text,
                excel_type=(c.data_type or "empty") if c.value is not None else "empty",
                fill_rgb=fill,
                hyperlink=(c.hyperlink.target if c.hyperlink else None),
                is_error=is_err,
            )
            grid.rows.setdefault(c.row, {})[c.column] = cell
        if row_has_content:
            last_content_row = row_obj[0].row if row_obj else last_content_row

    # max_row real = ultima fila con VALOR, no con formato
    grid.max_row = last_content_row

    # ---- celdas combinadas: fill-down explicito y auditado
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row > grid.max_row:
            continue
        anchor = grid.cell(rng.min_row, rng.min_col)
        if anchor is None or anchor.is_empty:
            continue
        for r in range(rng.min_row, min(rng.max_row, grid.max_row) + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                if r == rng.min_row and col == rng.min_col:
                    continue
                grid.rows.setdefault(r, {})[col] = Cell(
                    ref=CellRef(doc, sheet_name, r, col, get_column_letter(col)),
                    raw=anchor.raw,
                    text=anchor.text,
                    excel_type=anchor.excel_type,
                    fill_rgb=anchor.fill_rgb,
                    hyperlink=anchor.hyperlink,
                    inherited_from_row=rng.min_row,
                )
                grid.merged_applied += 1

    wb.close()
    return grid, anomalies

def list_sheets(path: Path, only_visible: bool = False) -> list[str]:
    """
    Lista hojas. Con only_visible=True omite las ocultas.
    Las hojas ocultas de una TS pertenecen a Digital, no a AdOps.
    """
    wb = load_workbook(path, data_only=True, keep_vba=False, read_only=True)
    if only_visible:
        names = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    else:
        names = list(wb.sheetnames)
    wb.close()
    return names

def sheet_states(path: Path) -> dict[str, str]:
    """{nombre: 'visible' | 'hidden' | 'veryHidden'}"""
    wb = load_workbook(path, data_only=True, keep_vba=False, read_only=True)
    out = {ws.title: ws.sheet_state for ws in wb.worksheets}
    wb.close()
    return out

# ---------------------------------------------------------------- header anchor

@dataclass
class HeaderResult:
    row: int | None = None
    mode: str = ""
    matches: int = 0
    evidence: str = ""

def find_header_row(grid: SheetGrid, spec: SheetSpec) -> tuple[HeaderResult, list[Anomaly]]:
    anomalies: list[Anomaly] = []
    all_keys: set[str] = set()
    for f in spec.fields:
        all_keys |= f.keys()

    # modo contractual: Innovid declara el header en la fila 10
    if spec.header_mode == "fixed_row" and spec.header_row:
        row = spec.header_row
        keys = {norm_key(t) for t in grid.row_text(row)}
        hits = len(keys & all_keys)
        if hits < spec.signature_min_matches:
            anomalies.append(Anomaly(
                "EXT-HEADER-FIXED-FAIL", "FATAL",
                f"Se esperaba el header en la fila {row} pero solo coinciden {hits} "
                f"campos conocidos (minimo {spec.signature_min_matches}). "
                f"El archivo pudo ser modificado sobre la fila 10.",
                detail={"row": row, "found": sorted(keys)[:15]},
            ))
            return HeaderResult(None, "fixed_row", hits), anomalies
        return HeaderResult(row, "fixed_row", hits,
                            f"fila fija {row}, {hits} campos"), anomalies

    # modo firma: se busca la fila con mas headers conocidos
    best_row, best_hits = None, 0
    for row in range(1, min(spec.signature_scan_rows, grid.max_row) + 1):
        keys = {norm_key(t) for t in grid.row_text(row)}
        hits = len(keys & all_keys)
        if hits > best_hits:
            best_row, best_hits = row, hits

    if best_row is None or best_hits < spec.signature_min_matches:
        anomalies.append(Anomaly(
            "EXT-HEADER-NOT-FOUND", "FATAL",
            f"No se encontro fila de header en las primeras {spec.signature_scan_rows} "
            f"filas (mejor candidata: fila {best_row} con {best_hits} coincidencias).",
        ))
        return HeaderResult(None, "signature", best_hits), anomalies

    return HeaderResult(best_row, "signature", best_hits,
                        f"fila {best_row}, {best_hits} campos"), anomalies

# ---------------------------------------------------------------- column mapper

@dataclass
class ColumnMap:
    single: dict[str, int] = field(default_factory=dict)
    multi: dict[str, list[int]] = field(default_factory=dict)
    unmapped: list[tuple[int, str]] = field(default_factory=list)
    missing_required: list[str] = field(default_factory=list)
    missing_optional: list[str] = field(default_factory=list)

    def has(self, name: str) -> bool:
        return name in self.single or bool(self.multi.get(name))

    def col(self, name: str) -> int | None:
        return self.single.get(name)

def map_columns(grid: SheetGrid, header_row: int,
                spec: SheetSpec) -> tuple[ColumnMap, list[Anomaly]]:
    """
    Mapea columnas POR NOMBRE. Nunca por posicion.
    Esta funcion elimina definitivamente los bugs B1 y B2.
    """
    anomalies: list[Anomaly] = []
    cmap = ColumnMap()

    key_to_field: dict[str, FieldSpec] = {}
    collisions: list[str] = []
    for f in spec.fields:
        for k in f.keys():
            prev = key_to_field.get(k)
            if prev is not None and prev.name != f.name:
                collisions.append(f"'{k}': {prev.name} vs {f.name}")
            key_to_field[k] = f

    if collisions:
        anomalies.append(Anomaly(
            "EXT-ALIAS-COLLISION", "FATAL",
            f"Dos campos canonicos comparten la misma llave normalizada: "
            f"{'; '.join(collisions)}. El mapeo seria impredecible.",
            detail={"collisions": collisions},
        ))

    header_cells = grid.row_cells(header_row)
    for col in sorted(header_cells):
        raw_header = header_cells[col].text
        if not raw_header:
            continue
        key = norm_key(raw_header)
        fld = key_to_field.get(key)

        if fld is None and key and key[-1].isdigit():
            # familias: Clicktag_1..20, Third_Party_Impression_1..10
            base = key.rstrip("0123456789")
            fld = key_to_field.get(base)
            if fld is not None and not fld.multi:
                fld = None

        if fld is None:
            cmap.unmapped.append((col, raw_header))
            continue

        if fld.multi:
            cmap.multi.setdefault(fld.name, []).append(col)
        elif fld.name not in cmap.single:
            cmap.single[fld.name] = col

    for f in spec.fields:
        if cmap.has(f.name):
            continue
        (cmap.missing_required if f.required else cmap.missing_optional).append(f.name)

    if cmap.missing_required:
        anomalies.append(Anomaly(
            "EXT-COLUMN-MISSING", "FATAL",
            f"Columnas requeridas ausentes: {', '.join(cmap.missing_required)}",
            detail={"missing": cmap.missing_required},
        ))

    return cmap, anomalies

def check_empty_required(grid: SheetGrid, cmap: ColumnMap, spec: SheetSpec,
                         first_data_row: int, last_row: int) -> list[Anomaly]:
    """
    Regla EXT-001: una columna requerida presente pero 100% vacia es una
    ANOMALIA DE EXTRACCION, no un hallazgo de QA. Aborta el run.

    Salvaguarda contra el estado original del VBA: reportar numeros
    sin haber leido nada.
    """
    out: list[Anomaly] = []
    for f in spec.fields:
        if not f.required or f.multi:
            continue
        col = cmap.col(f.name)
        if col is None:
            continue
        filled = sum(
            1 for r in range(first_data_row, last_row + 1)
            if (c := grid.cell(r, col)) is not None and c.has_value
        )
        if filled == 0:
            out.append(Anomaly(
                "EXT-001", "FATAL",
                f"La columna requerida '{f.name}' existe pero esta 100% vacia. "
                f"Probable error de lectura o de exportacion. Se aborta.",
                detail={"field": f.name, "column": col},
            ))
    return out

def column_fill_rates(grid: SheetGrid, cmap: ColumnMap, spec: SheetSpec,
                      data_rows: list[int]) -> dict[str, float]:
    """
    % de celdas con valor real por campo canonico.

    Una columna presente pero 0% poblada NO habilita su dominio de reglas.
    Sin esto, el Capability Profile dice 'puedo validar X' cuando no puede.
    """
    if not data_rows:
        return {}
    total = len(data_rows)
    out: dict[str, float] = {}

    for name, col in cmap.single.items():
        filled = sum(
            1 for r in data_rows
            if (c := grid.cell(r, col)) is not None and c.has_value
        )
        out[name] = filled / total

    for name, cols in cmap.multi.items():
        filled = sum(
            1 for r in data_rows
            if any((c := grid.cell(r, col)) is not None and c.has_value for col in cols)
        )
        out[name] = filled / total

    return out

# ---------------------------------------------------------------- block segmenter

def classify_row(grid: SheetGrid, row: int, cmap: ColumnMap, spec: SheetSpec,
                 primary_key: str) -> tuple[str, str]:
    """
    Clasifica una fila ANTES de convertirla en dato.
    Devuelve (clase, motivo).

    Orden optimizado: primero los chequeos baratos, los regex al final
    y solo sobre las primeras celdas de la fila.
    """
    cells = grid.row_cells(row)
    if not cells:
        return ROW_SEPARATOR, "fila vacia"

    texts = [c.text for c in cells.values() if c.text]
    if not texts:
        return ROW_SEPARATOR, "fila vacia"

    # --- barato: errores de formula
    for c in cells.values():
        if c.is_error:
            return ROW_CORRUPT, "contiene error de formula"

    # --- barato: marcadores de corrupcion
    for t in texts[:12]:
        if "[object object]" in t.casefold():
            return ROW_CORRUPT, "contiene [object Object]"

    # --- barato: banner (un solo valor repetido a lo ancho)
    if len(texts) >= 3 and len({t.casefold() for t in texts}) == 1:
        return ROW_BANNER, f"un solo valor repetido: '{texts[0][:40]}'"

    # --- caro: patrones de plantilla, solo sobre las primeras celdas
    if spec.ignore_row_patterns:
        head = texts[:4]
        for pat in spec.ignore_row_patterns:
            for t in head:
                if re.search(pat, t, re.IGNORECASE):
                    return ROW_TEMPLATE, f"coincide patron '{pat}'"

    # --- llave primaria
    pk_col = cmap.col(primary_key)
    if pk_col is not None:
        pk_cell = grid.cell(row, pk_col)
        if pk_cell is None or not pk_cell.has_value:
            if len(texts) <= 2:
                return ROW_BANNER, f"sin {primary_key} y casi vacia"
            return ROW_DATA, f"sin {primary_key} (posible fila de cabecera)"

    # --- identidad de entidad (transiciones del export)
    if spec.entity_identity_any:
        applies = True
        if spec.entity_identity_when:
            applies = False
            for fname in spec.entity_identity_when:
                col = cmap.col(fname)
                if col is None:
                    continue
                c = grid.cell(row, col)
                if c is not None and c.has_value:
                    applies = True
                    break

        if applies:
            has_any = False
            checked = False
            for fname in spec.entity_identity_any:
                col = cmap.col(fname)
                if col is None:
                    continue
                checked = True
                c = grid.cell(row, col)
                if c is not None and c.has_value:
                    has_any = True
                    break
            if checked and not has_any:
                return ROW_SEPARATOR, (
                    "con " + "/".join(spec.entity_identity_when) +
                    " pero sin " + "/".join(spec.entity_identity_any) +
                    " (transicion entre bloques)"
                )

    return ROW_DATA, ""
def resolve_sheet(available: list[str], aliases: list[str]) -> str | None:
    """
    Encuentra la hoja por alias. Match exacto normalizado primero,
    luego 'contiene'. Devuelve None si no hay candidata.
    """
    norm_avail = {norm_key(s): s for s in available}
    for a in aliases:
        hit = norm_avail.get(norm_key(a))
        if hit:
            return hit
    for a in aliases:
        ka = norm_key(a)
        for k, orig in norm_avail.items():
            if ka and ka in k:
                return orig
    return None